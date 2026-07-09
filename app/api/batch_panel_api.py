"""Batch report panel API: edit a site config YAML and run the whole
extraction (gate copy → crossings → merged workbooks → client template)
from the viewer UI. Reuses batch_report.run_from_config — identical
behavior to the CLI."""

import os
import sys
import threading

from fastapi import APIRouter
from fastapi.responses import FileResponse

# batch_report.py lives in the app root (same folder as run.py)
_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

router = APIRouter(tags=["batch"])

# Single in-process run state (desktop app: one run at a time)
STATE = {'running': False, 'log': [], 'outputs': [], 'warnings': [],
         'error': None, 'done': False}


@router.get("/batch/config")
def load_config(path: str):
    """Load a site YAML (returns {} if the file doesn't exist yet)."""
    try:
        import yaml
        path = path.strip().strip('"\'').strip()
        if not os.path.exists(path):
            return {'config': {}, 'exists': False}
        with open(path, encoding='utf-8') as f:
            cfg = yaml.safe_load(f) or {}
        return {'config': cfg, 'exists': True}
    except Exception as e:
        return {'error': str(e)}


@router.post("/batch/config")
def save_config(payload: dict):
    """Save a site YAML. payload: {path, config}"""
    try:
        import yaml
        path = str(payload['path']).strip().strip('"\'').strip()
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        with open(path, 'w', encoding='utf-8') as f:
            yaml.safe_dump(payload.get('config') or {}, f,
                           sort_keys=False, allow_unicode=True)
        return {'saved': path}
    except Exception as e:
        return {'error': str(e)}


@router.get("/batch/template_info")
def template_info(path: str):
    """Read a client template: movement tab numbers + class column headers.
    Used to populate the mapping dropdowns."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        tabs = [n.replace('Movement ', '').strip()
                for n in wb.sheetnames if n.startswith('Movement ')
                and n != 'Movement Generator']
        headers = []
        for n in wb.sheetnames:
            if n.startswith('Movement ') and n != 'Movement Generator':
                ws = wb[n]
                for row in ws.iter_rows(min_row=1, max_row=60, values_only=True):
                    if row and str(row[0]).strip().upper() == 'TIME':
                        headers = [str(c).strip() for c in row[1:]
                                   if c and str(c).strip().upper() != 'TOTAL']
                        break
                if headers:
                    break
        wb.close()
        return {'movement_tabs': tabs, 'class_headers': headers}
    except Exception as e:
        return {'error': str(e)}


@router.get("/batch/our_classes")
def our_classes():
    """Distinct vehicle classes in the currently loaded .traf."""
    try:
        from app.core.database import query
        rows = query("SELECT DISTINCT class_name FROM tracks "
                     "WHERE class_name IS NOT NULL ORDER BY class_name")
        return {'classes': [r['class_name'] for r in rows]}
    except Exception as e:
        return {'error': str(e)}


@router.get("/batch/current_traf")
def current_traf():
    """Path of the currently loaded .traf (handy default for gates_from)."""
    try:
        from app.core.database import state
        return {'path': state.get('db_path')}
    except Exception:
        return {'path': None}


def _run_thread(cfg):
    try:
        import batch_report
        result = batch_report.run_from_config(
            cfg, log=lambda msg: STATE['log'].append(str(msg)))
        STATE['outputs'] = result['outputs']
        STATE['warnings'] = result['warnings']
    except Exception as e:
        STATE['error'] = str(e)
        STATE['log'].append(f"ERROR: {e}")
    finally:
        STATE['running'] = False
        STATE['done'] = True


@router.post("/batch/run")
def run_batch(payload: dict):
    """Start a batch run. payload: {config: {...}} (the site YAML content)."""
    if STATE['running']:
        return {'error': 'A batch run is already in progress'}
    STATE.update({'running': True, 'log': [], 'outputs': [], 'warnings': [],
                  'error': None, 'done': False})
    t = threading.Thread(target=_run_thread, args=(payload.get('config') or {},),
                         daemon=True)
    t.start()
    return {'started': True}


@router.get("/batch/status")
def status():
    return {'running': STATE['running'], 'done': STATE['done'],
            'error': STATE['error'], 'log': STATE['log'],
            'outputs': STATE['outputs'], 'warnings': STATE['warnings']}


@router.get("/batch/download")
def download(path: str):
    """Download an output file produced by the last run (restricted to
    files the run actually produced, to avoid arbitrary file reads)."""
    real = os.path.abspath(path)
    if real not in [os.path.abspath(p) for p in STATE['outputs']]:
        return {'error': 'Not an output of the last run'}
    return FileResponse(real, filename=os.path.basename(real))
