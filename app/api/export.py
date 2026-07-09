"""
Export gate crossing data to Excel.
"""
import io
from datetime import datetime
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from app.core.database import query, state, get_conn
from app.core.movement_inference import get_movement_assignments

router = APIRouter(tags=["export"])

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@router.get("/export/gates_excel")
def export_gates_excel():
    """
    Export gate crossing counts as a formatted Excel file.
    Includes: summary sheet, per-gate sheets, and OD matrix.
    """
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    gates = query("SELECT * FROM gates")
    if not gates:
        raise HTTPException(400, "No gates defined. Draw gates first.")

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    sub_header_fill = PatternFill('solid', fgColor='D6E4F0')
    sub_header_font = Font(name='Arial', bold=True, size=10)
    data_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', bold=True, size=10, color='2F5496')
    total_fill = PatternFill('solid', fgColor='E2EFDA')
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'))

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def style_data(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # ═══════════════════════════════════════════
    # SHEET 1: Summary
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = 'Summary'

    # Title
    ws['A1'] = 'Traffic Gate Crossing Report'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=9, color='808080')
    ws['A3'] = f'Video: {state.get("db_path", "Unknown")}'
    ws['A3'].font = Font(name='Arial', size=9, color='808080')

    # Gate summary table
    row = 5
    from app.core.class_profile import get_class_profile
    profile = get_class_profile(get_conn())
    vehicle_classes = profile['vehicle_classes']
    headers = ['Gate Name', 'Direction', 'Total Count'] + vehicle_classes + ['Other']
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, len(headers))
    row += 1

    for g in gates:
        crossings = query(
            "SELECT gc.class_name, COUNT(*) as cnt "
            "FROM gate_crossings gc WHERE gc.gate_id=? "
            "GROUP BY gc.class_name", (g['gate_id'],))
        counts = {c['class_name']: c['cnt'] for c in crossings}
        total = sum(counts.values())

        ws.cell(row=row, column=1, value=g['name'])
        ws.cell(row=row, column=2, value=g.get('direction', 'both'))
        ws.cell(row=row, column=3, value=total)
        for ci, cls in enumerate(vehicle_classes):
            ws.cell(row=row, column=4 + ci, value=counts.get(cls, 0))
        other = total - sum(counts.get(c, 0) for c in vehicle_classes)
        ws.cell(row=row, column=4 + len(vehicle_classes), value=other)
        style_data(ws, row, len(headers))
        row += 1

    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value='TOTAL')
    for c in range(3, len(headers) + 1):
        col_letter = chr(64 + c)
        ws.cell(row=total_row, column=c,
                value=f'=SUM({col_letter}6:{col_letter}{total_row - 1})')
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for col in 'CDEFGHI':
        ws.column_dimensions[col].width = 10

    # ═══════════════════════════════════════════
    # SHEET 2: Origin-Destination Matrix
    # ═══════════════════════════════════════════
    ws_od = wb.create_sheet('OD Matrix')

    ws_od['A1'] = 'Origin-Destination Matrix (Gate to Gate)'
    ws_od['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_od['A2'] = 'Entry gate (rows) → Exit gate (columns)'
    ws_od['A2'].font = Font(name='Arial', size=9, color='808080')

    gate_names = [g['name'] for g in gates]
    gate_ids = [g['gate_id'] for g in gates]

    # Headers
    row = 4
    ws_od.cell(row=row, column=1, value='From \\ To')
    for c, name in enumerate(gate_names, 2):
        ws_od.cell(row=row, column=c, value=name)
    style_header(ws_od, row, len(gate_names) + 1)

    # For each pair of gates, find tracks that crossed both
    for ri, (from_id, from_name) in enumerate(zip(gate_ids, gate_names)):
        r = row + 1 + ri
        ws_od.cell(row=r, column=1, value=from_name)
        ws_od.cell(row=r, column=1).font = sub_header_font
        ws_od.cell(row=r, column=1).fill = sub_header_fill
        ws_od.cell(row=r, column=1).border = thin_border

        for ci, (to_id, to_name) in enumerate(zip(gate_ids, gate_names)):
            if from_id == to_id:
                ws_od.cell(row=r, column=ci + 2, value='-')
            else:
                # Count tracks that crossed from_gate then to_gate
                result = query(
                    "SELECT COUNT(DISTINCT gc1.track_id) as cnt "
                    "FROM gate_crossings gc1 "
                    "JOIN gate_crossings gc2 ON gc1.track_id = gc2.track_id "
                    "WHERE gc1.gate_id=? AND gc2.gate_id=? AND gc1.frame < gc2.frame",
                    (from_id, to_id))
                cnt = result[0]['cnt'] if result else 0
                ws_od.cell(row=r, column=ci + 2, value=cnt)
            style_data(ws_od, r, len(gate_names) + 1)

    ws_od.column_dimensions['A'].width = 18
    for i, _ in enumerate(gate_names):
        ws_od.column_dimensions[chr(66 + i)].width = 15

    # ═══════════════════════════════════════════
    # SHEET 3: Detailed Crossings
    # ═══════════════════════════════════════════
    ws_det = wb.create_sheet('Detailed Crossings')

    ws_det['A1'] = 'Detailed Gate Crossing Log'
    ws_det['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

    row = 3
    det_headers = ['Gate', 'Track ID', 'Class', 'Frame', 'Direction', 'Speed (px/f)']
    for c, h in enumerate(det_headers, 1):
        ws_det.cell(row=row, column=c, value=h)
    style_header(ws_det, row, len(det_headers))
    row += 1

    all_crossings = query(
        "SELECT gc.*, g.name as gate_name "
        "FROM gate_crossings gc JOIN gates g ON gc.gate_id = g.gate_id "
        "ORDER BY g.name, gc.frame")

    for cr in all_crossings[:5000]:  # Cap at 5000 rows
        ws_det.cell(row=row, column=1, value=cr['gate_name'])
        ws_det.cell(row=row, column=2, value=cr['track_id'])
        ws_det.cell(row=row, column=3, value=cr.get('class_name', ''))
        ws_det.cell(row=row, column=4, value=cr['frame'])
        ws_det.cell(row=row, column=5, value=cr.get('direction', ''))
        ws_det.cell(row=row, column=6, value=round(cr.get('speed_px', 0), 1))
        style_data(ws_det, row, len(det_headers))
        row += 1

    ws_det.column_dimensions['A'].width = 18
    ws_det.column_dimensions['B'].width = 10
    ws_det.column_dimensions['C'].width = 10
    ws_det.column_dimensions['D'].width = 10
    ws_det.column_dimensions['E'].width = 12
    ws_det.column_dimensions['F'].width = 14

    # ═══════════════════════════════════════════
    # SHEET 4: Per-class breakdown per gate
    # ═══════════════════════════════════════════
    ws_cls = wb.create_sheet('Class Breakdown')

    ws_cls['A1'] = 'Vehicle Classification per Gate'
    ws_cls['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

    row = 3
    for g in gates:
        ws_cls.cell(row=row, column=1, value=g['name'])
        ws_cls.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=11)
        ws_cls.cell(row=row, column=1).fill = sub_header_fill
        row += 1

        cls_headers = ['Class', 'Count', 'Percentage']
        for c, h in enumerate(cls_headers, 1):
            ws_cls.cell(row=row, column=c, value=h)
        style_header(ws_cls, row, len(cls_headers))
        row += 1

        crossings = query(
            "SELECT gc.class_name, COUNT(*) as cnt "
            "FROM gate_crossings gc WHERE gc.gate_id=? "
            "GROUP BY gc.class_name ORDER BY cnt DESC",
            (g['gate_id'],))
        total = sum(c['cnt'] for c in crossings)

        start_row = row
        for cr in crossings:
            ws_cls.cell(row=row, column=1, value=cr['class_name'])
            ws_cls.cell(row=row, column=2, value=cr['cnt'])
            pct = (cr['cnt'] / total * 100) if total > 0 else 0
            ws_cls.cell(row=row, column=3, value=round(pct, 1))
            ws_cls.cell(row=row, column=3).number_format = '0.0"%"'
            style_data(ws_cls, row, len(cls_headers))
            row += 1

        # Total for this gate
        ws_cls.cell(row=row, column=1, value='Total')
        ws_cls.cell(row=row, column=2, value=f'=SUM(B{start_row}:B{row - 1})')
        ws_cls.cell(row=row, column=3, value=100.0)
        for c in range(1, 4):
            ws_cls.cell(row=row, column=c).font = total_font
            ws_cls.cell(row=row, column=c).fill = total_fill
            ws_cls.cell(row=row, column=c).border = thin_border
        row += 2

    ws_cls.column_dimensions['A'].width = 15
    ws_cls.column_dimensions['B'].width = 10
    ws_cls.column_dimensions['C'].width = 12

    # ── Save to buffer ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'gate_crossings_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ═══════════════════════════════════════════════════════════
# MOVEMENT-BASED EXPORT
# ═══════════════════════════════════════════════════════════

def _write_inference_report_sheet(wb, report, style_cell, header_font,
                                  header_fill, data_font, total_font,
                                  total_fill):
    """
    Shared 'Inference Report' sheet: per-movement Direct/Inferred counts,
    the learned per-site signatures (direction signs + mean headings),
    and zero-crossing suggestions. This is the QA trail when running
    hundreds of sites in batch.
    """
    ws = wb.create_sheet('Inference Report')
    ws['A1'] = 'Movement Inference Report'
    ws['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws['A2'] = ('Reference paths are learned per movement from its direct tracks '
                '(median trajectory). Partial tracks are assigned when they lie '
                'within the learned corridor with matching direction. Zero-crossing '
                'tracks are counted only on a strict, unambiguous shape match.')
    ws['A2'].font = Font(name='Arial', size=9, color='808080')

    row = 4
    hdrs = ['Movement', 'Direct', 'Inferred', 'Total',
            'FROM-gate sign (agree)', 'TO-gate sign (agree)',
            'Ref Tracks', 'Corridor (px)', 'Reference Trusted']
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=row, column=c, value=h)
        style_cell(ws.cell(row=row, column=c), font=header_font, fill=header_fill)
    row += 1
    data_start = row

    for m in report['per_movement']:
        vals = [m['label'], m['direct'], m['inferred'], m['total'],
                f"{m['from_sign'] or '-'} ({m['from_sign_agree']})",
                f"{m['to_sign'] or '-'} ({m['to_sign_agree']})",
                m['ref_tracks'],
                m['corridor_px'] if m['corridor_px'] is not None else '-',
                'yes' if m['ref_tracks'] >= report['params']['min_confirmed_for_reference']
                else f"no (needs >= {report['params']['min_confirmed_for_reference']} direct)"]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
            style_cell(ws.cell(row=row, column=c),
                       align='left' if c in (1, 5, 6, 9) else 'center')
        row += 1

    # Totals
    ws.cell(row=row, column=1, value='TOTAL')
    ws.cell(row=row, column=2, value=report['total_direct'])
    ws.cell(row=row, column=3, value=report['total_inferred'])
    ws.cell(row=row, column=4, value=report['total_direct'] + report['total_inferred'])
    for c in range(1, 5):
        style_cell(ws.cell(row=row, column=c), font=total_font, fill=total_fill,
                   align='left' if c == 1 else 'center')
    row += 1
    ws.cell(row=row, column=1,
            value=f"Of the inferred, recovered with zero crossings: "
                  f"{report.get('inferred_zero_crossing', 0)}")
    ws.cell(row=row, column=1).font = Font(name='Arial', size=10, color='B45309')
    row += 1
    ws.cell(row=row, column=1,
            value=f"Unresolved (see Unassigned Tracks sheet): {report['unresolved_count']}")
    ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=10, color='C00000')
    row += 2

    widths = [24, 10, 10, 10, 22, 22, 11, 13, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _resolve_movements(movements_json):
    """
    Parse the movements query param, or auto-generate all ordered
    leg-gate pairs (including U-turns) when it is empty/'auto'.
    Also ensures gate crossings were computed by engine v3 (which
    records multiple crossings per gate — required for U-turns);
    older stored crossings are recomputed once.
    """
    import json as json_lib
    from app.core.gate_engine import recompute_all_gates, ENGINE_VERSION
    from app.core.movement_inference import auto_generate_movements

    conn = get_conn()
    ver = None
    try:
        row = conn.execute(
            "SELECT value FROM scene WHERE key='crossings_engine_version'").fetchone()
        ver = row[0] if row else None
    except Exception:
        pass
    if ver != ENGINE_VERSION:
        recompute_all_gates(conn)

    raw = (movements_json or '').strip()
    if raw in ('', '[]', 'auto', 'null'):
        movement_list = auto_generate_movements(conn, include_uturns=True)
        if not movement_list:
            raise HTTPException(400, "No gates defined. Draw one gate per leg/approach first.")
        return movement_list
    try:
        movement_list = json_lib.loads(raw)
    except Exception:
        raise HTTPException(400, "Invalid movements JSON")
    if not movement_list:
        raise HTTPException(400, "No movements defined")
    return movement_list


def _mov_display(from_name, to_name):
    label = f"{from_name} → {to_name}"
    if from_name == to_name:
        label += " (U-turn)"
    return label


@router.get("/export/movement_excel")
def export_movement_excel(movements: str = ''):
    """
    Export directional movement counts as Excel.
    movements: JSON array [{from_id, from_name, to_id, to_name}, ...] or
    empty/'auto' to generate all ordered leg-gate pairs incl. U-turns.
    """
    import json as json_lib
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl not installed")

    movement_list = _resolve_movements(movements)

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    sub_header_fill = PatternFill('solid', fgColor='D6E4F0')
    sub_header_font = Font(name='Arial', bold=True, size=10)
    data_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', bold=True, size=10, color='2F5496')
    total_fill = PatternFill('solid', fgColor='E2EFDA')
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'))

    def style_cell(cell, font=data_font, fill=None, align='center'):
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align)
        if fill:
            cell.fill = fill

    # Get all class names in the data
    all_classes_raw = query(
        "SELECT DISTINCT gc.class_name FROM gate_crossings gc ORDER BY gc.class_name")
    class_names = [r['class_name'] for r in all_classes_raw if r['class_name']]
    if not class_names:
        try:
            from app.core.class_profile import get_class_profile
            profile = get_class_profile(get_conn())
            class_names = profile['vehicle_classes']
        except Exception:
            class_names = ['PV', 'SU', 'CU']

    # ── Direct + Inferred assignments (one pass for all movements) ──
    # Inferred = single-crossing tracks recovered via per-site learned
    # direction-sign / heading signatures (app/core/movement_inference.py).
    assign_data = get_movement_assignments(get_conn(), movement_list)
    assignments = assign_data['assignments']
    inference_report = assign_data['report']
    unresolved_tracks = assign_data['unresolved']

    # Track speeds (one query, reused for Movement Details)
    track_speed = {r['track_id']: (r['speed_mean_px'] or 0.0)
                   for r in query("SELECT track_id, speed_mean_px FROM tracks")}

    # ═══════════════════════════════════════
    # SHEET 1: Movement Summary
    # ═══════════════════════════════════════
    ws = wb.active
    ws.title = 'Movement Summary'

    ws['A1'] = 'Traffic Movement Report'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=9, color='808080')
    ws['A3'] = f'Source: {state.get("db_path", "Unknown")}'
    ws['A3'].font = Font(name='Arial', size=9, color='808080')
    ws['A4'] = ('Direct = crossed both gates. Inferred = crossed one gate; '
                'movement recovered from crossing direction + trajectory heading. '
                'See "Inference Report" sheet.')
    ws['A4'].font = Font(name='Arial', size=9, color='808080')

    # Headers
    row = 6
    headers = ['Movement', 'From Gate', 'To Gate',
               'Direct', 'Inferred', 'Total'] + class_names
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        style_cell(cell, font=header_font, fill=header_fill)

    # Each movement
    row = 7
    data_start_row = row
    for mov in movement_list:
        from_id = mov['from_id']
        to_id = mov['to_id']
        from_name = mov.get('from_name', str(from_id))
        to_name = mov.get('to_name', str(to_id))
        movement_label = _mov_display(from_name, to_name)

        rows_a = assignments.get((from_id, to_id), [])
        direct_cnt = sum(1 for a in rows_a if not a['inferred'])
        inferred_cnt = sum(1 for a in rows_a if a['inferred'])

        # Per-class totals (direct + inferred)
        class_counts = {cls: 0 for cls in class_names}
        for a in rows_a:
            if a['class_name'] in class_counts:
                class_counts[a['class_name']] += 1

        ws.cell(row=row, column=1, value=movement_label)
        style_cell(ws.cell(row=row, column=1), align='left')
        ws.cell(row=row, column=2, value=from_name)
        style_cell(ws.cell(row=row, column=2))
        ws.cell(row=row, column=3, value=to_name)
        style_cell(ws.cell(row=row, column=3))
        ws.cell(row=row, column=4, value=direct_cnt)
        style_cell(ws.cell(row=row, column=4))
        ws.cell(row=row, column=5, value=inferred_cnt)
        style_cell(ws.cell(row=row, column=5),
                   font=Font(name='Arial', size=10, color='B45309'))
        ws.cell(row=row, column=6, value=direct_cnt + inferred_cnt)
        style_cell(ws.cell(row=row, column=6), font=Font(name='Arial', bold=True, size=10))

        for ci, cls in enumerate(class_names):
            ws.cell(row=row, column=7 + ci, value=class_counts.get(cls, 0))
            style_cell(ws.cell(row=row, column=7 + ci))

        row += 1

    # Total row
    total_row = row
    ws.cell(row=total_row, column=1, value='TOTAL')
    style_cell(ws.cell(row=total_row, column=1), font=total_font, fill=total_fill, align='left')
    for c in range(2, 4):
        style_cell(ws.cell(row=total_row, column=c), font=total_font, fill=total_fill)
    for c in range(4, len(headers) + 1):
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c,
                value=f'=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})')
        style_cell(ws.cell(row=total_row, column=c), font=total_font, fill=total_fill)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    for c in range(4, 7):
        ws.column_dimensions[get_column_letter(c)].width = 10
    for i in range(len(class_names)):
        ws.column_dimensions[get_column_letter(7 + i)].width = 8

    # ═══════════════════════════════════════
    # SHEET 2: OD Matrix (all gates × all gates)
    # ═══════════════════════════════════════
    ws_od = wb.create_sheet('OD Matrix')

    ws_od['A1'] = 'Origin-Destination Matrix'
    ws_od['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_od['A2'] = 'Rows = Entry gate, Columns = Exit gate (direct crossings only)'
    ws_od['A2'].font = Font(name='Arial', size=9, color='808080')

    all_gates = query("SELECT * FROM gates")
    gate_names = [g['name'] for g in all_gates]
    gate_ids = [g['gate_id'] for g in all_gates]

    row = 4
    ws_od.cell(row=row, column=1, value='From \\ To')
    style_cell(ws_od.cell(row=row, column=1), font=header_font, fill=header_fill)
    # Add "Total" column after all gates
    for c, name in enumerate(gate_names, 2):
        ws_od.cell(row=row, column=c, value=name)
        style_cell(ws_od.cell(row=row, column=c), font=header_font, fill=header_fill)
    ws_od.cell(row=row, column=len(gate_names) + 2, value='Total')
    style_cell(ws_od.cell(row=row, column=len(gate_names) + 2), font=header_font, fill=header_fill)

    for ri, (from_id, from_name) in enumerate(zip(gate_ids, gate_names)):
        r = row + 1 + ri
        ws_od.cell(row=r, column=1, value=from_name)
        style_cell(ws_od.cell(row=r, column=1), font=sub_header_font, fill=sub_header_fill, align='left')

        for ci, (to_id, to_name) in enumerate(zip(gate_ids, gate_names)):
            if from_id == to_id:
                ws_od.cell(row=r, column=ci + 2, value='-')
            else:
                result = query(
                    "SELECT COUNT(DISTINCT gc1.track_id) as cnt "
                    "FROM gate_crossings gc1 "
                    "JOIN gate_crossings gc2 ON gc1.track_id = gc2.track_id "
                    "WHERE gc1.gate_id=? AND gc2.gate_id=? AND gc1.frame < gc2.frame",
                    (from_id, to_id))
                cnt = result[0]['cnt'] if result else 0
                ws_od.cell(row=r, column=ci + 2, value=cnt)
            style_cell(ws_od.cell(row=r, column=ci + 2))

        # Row total formula
        first_col = get_column_letter(2)
        last_col = get_column_letter(1 + len(gate_names))
        ws_od.cell(row=r, column=len(gate_names) + 2,
                   value=f'=SUM({first_col}{r}:{last_col}{r})')
        style_cell(ws_od.cell(row=r, column=len(gate_names) + 2), font=total_font, fill=total_fill)

    ws_od.column_dimensions['A'].width = 18
    for i in range(len(gate_names) + 1):
        ws_od.column_dimensions[get_column_letter(2 + i)].width = 14

    # ═══════════════════════════════════════
    # SHEET 3: Per-Movement Detail (class breakdown)
    # ═══════════════════════════════════════
    ws_det = wb.create_sheet('Movement Details')

    ws_det['A1'] = 'Movement Details — Per-Class Breakdown'
    ws_det['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

    row = 3
    for mov in movement_list:
        from_id = mov['from_id']
        to_id = mov['to_id']
        from_name = mov.get('from_name', str(from_id))
        to_name = mov.get('to_name', str(to_id))

        ws_det.cell(row=row, column=1, value=_mov_display(from_name, to_name))
        style_cell(ws_det.cell(row=row, column=1), font=Font(name='Arial', bold=True, size=11), fill=sub_header_fill, align='left')
        ws_det.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        det_headers = ['Class', 'Direct', 'Inferred', 'Total', 'Avg Speed (px/f)']
        for c, h in enumerate(det_headers, 1):
            ws_det.cell(row=row, column=c, value=h)
            style_cell(ws_det.cell(row=row, column=c), font=header_font, fill=header_fill)
        row += 1

        rows_a = assignments.get((from_id, to_id), [])
        start_row = row
        for cls in class_names:
            cls_rows = [a for a in rows_a if a['class_name'] == cls]
            if not cls_rows:
                continue
            direct_n = sum(1 for a in cls_rows if not a['inferred'])
            inferred_n = len(cls_rows) - direct_n
            speeds = [track_speed.get(a['track_id'], 0.0) for a in cls_rows]
            avg_spd = sum(speeds) / len(speeds) if speeds else 0

            ws_det.cell(row=row, column=1, value=cls)
            style_cell(ws_det.cell(row=row, column=1))
            ws_det.cell(row=row, column=2, value=direct_n)
            style_cell(ws_det.cell(row=row, column=2))
            ws_det.cell(row=row, column=3, value=inferred_n)
            style_cell(ws_det.cell(row=row, column=3),
                       font=Font(name='Arial', size=10, color='B45309'))
            ws_det.cell(row=row, column=4, value=len(cls_rows))
            style_cell(ws_det.cell(row=row, column=4),
                       font=Font(name='Arial', bold=True, size=10))
            ws_det.cell(row=row, column=5, value=round(avg_spd, 1))
            style_cell(ws_det.cell(row=row, column=5))
            row += 1

        # Total for this movement
        if row > start_row:
            ws_det.cell(row=row, column=1, value='Total')
            for c in range(2, 5):
                col_letter = get_column_letter(c)
                ws_det.cell(row=row, column=c,
                            value=f'=SUM({col_letter}{start_row}:{col_letter}{row - 1})')
            for c in range(1, 6):
                style_cell(ws_det.cell(row=row, column=c), font=total_font, fill=total_fill)
        row += 2  # gap before next movement

    ws_det.column_dimensions['A'].width = 12
    ws_det.column_dimensions['B'].width = 10
    ws_det.column_dimensions['C'].width = 10
    ws_det.column_dimensions['D'].width = 10
    ws_det.column_dimensions['E'].width = 18

    # ═══════════════════════════════════════
    # SHEET 4: Gate Summary (per gate totals)
    # ═══════════════════════════════════════
    ws_gate = wb.create_sheet('Gate Totals')

    ws_gate['A1'] = 'Per-Gate Crossing Totals'
    ws_gate['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')

    row = 3
    gate_headers = ['Gate', 'Total', 'Direction'] + class_names
    for c, h in enumerate(gate_headers, 1):
        ws_gate.cell(row=row, column=c, value=h)
        style_cell(ws_gate.cell(row=row, column=c), font=header_font, fill=header_fill)
    row += 1

    for g in all_gates:
        crossings = query(
            "SELECT gc.class_name, COUNT(*) as cnt "
            "FROM gate_crossings gc WHERE gc.gate_id=? "
            "GROUP BY gc.class_name", (g['gate_id'],))
        counts = {c['class_name']: c['cnt'] for c in crossings}
        total = sum(counts.values())

        ws_gate.cell(row=row, column=1, value=g['name'])
        style_cell(ws_gate.cell(row=row, column=1), align='left')
        ws_gate.cell(row=row, column=2, value=total)
        style_cell(ws_gate.cell(row=row, column=2), font=Font(name='Arial', bold=True, size=10))
        ws_gate.cell(row=row, column=3, value=g.get('direction', 'both'))
        style_cell(ws_gate.cell(row=row, column=3))
        for ci, cls in enumerate(class_names):
            ws_gate.cell(row=row, column=4 + ci, value=counts.get(cls, 0))
            style_cell(ws_gate.cell(row=row, column=4 + ci))
        row += 1

    ws_gate.column_dimensions['A'].width = 18
    ws_gate.column_dimensions['B'].width = 10
    ws_gate.column_dimensions['C'].width = 12
    for i in range(len(class_names)):
        ws_gate.column_dimensions[chr(68 + i)].width = 8

    # ═══════════════════════════════════════
    # SHEET 5: Track IDs per Movement
    # ═══════════════════════════════════════
    ws_ids = wb.create_sheet('Track IDs by Movement')

    ws_ids['A1'] = 'Track IDs per Movement'
    ws_ids['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_ids['A2'] = 'Each column lists the track IDs that crossed from one gate to another'
    ws_ids['A2'].font = Font(name='Arial', size=9, color='808080')

    col = 1
    all_assigned_ids = set()

    for mov in movement_list:
        from_id = mov['from_id']
        to_id = mov['to_id']
        from_name = mov.get('from_name', str(from_id))
        to_name = mov.get('to_name', str(to_id))
        movement_label = _mov_display(from_name, to_name).replace('→', '->')

        # All assigned tracks (direct + inferred), ordered by entry frame
        id_rows = sorted(assignments.get((from_id, to_id), []),
                         key=lambda a: a['entry_frame'])

        # Header row
        hdrs = [movement_label, 'Class', 'Entry Frame', 'Exit Frame', 'Assigned By']
        for ci, h in enumerate(hdrs):
            cell = ws_ids.cell(row=4, column=col + ci, value=h)
            style_cell(cell, font=header_font, fill=header_fill,
                       align='left' if ci == 0 else 'center')

        # Count header
        n_direct = sum(1 for a in id_rows if not a['inferred'])
        n_inferred = len(id_rows) - n_direct
        ws_ids.cell(row=3, column=col,
                    value=f'Count: {len(id_rows)} ({n_direct} direct + {n_inferred} inferred)')
        ws_ids.cell(row=3, column=col).font = Font(name='Arial', bold=True, size=10, color='2F5496')

        for ri, a in enumerate(id_rows):
            ws_ids.cell(row=5 + ri, column=col, value=a['track_id'])
            style_cell(ws_ids.cell(row=5 + ri, column=col), align='left')
            ws_ids.cell(row=5 + ri, column=col + 1, value=a['class_name'])
            style_cell(ws_ids.cell(row=5 + ri, column=col + 1))
            ws_ids.cell(row=5 + ri, column=col + 2, value=a['entry_frame'])
            style_cell(ws_ids.cell(row=5 + ri, column=col + 2))
            ws_ids.cell(row=5 + ri, column=col + 3,
                        value=a['exit_frame'] if a['exit_frame'] is not None else '-')
            style_cell(ws_ids.cell(row=5 + ri, column=col + 3))
            basis = 'Direct' if not a['inferred'] else f"Inferred: {a['basis']}"
            cell = ws_ids.cell(row=5 + ri, column=col + 4, value=basis)
            style_cell(cell, align='left',
                       font=(data_font if not a['inferred']
                             else Font(name='Arial', size=10, color='B45309')))
            all_assigned_ids.add(a['track_id'])

        # Set column widths
        for wi, w in enumerate([22, 8, 12, 12, 22]):
            ws_ids.column_dimensions[get_column_letter(col + wi)].width = w

        col += 6  # gap between movements

    # ═══════════════════════════════════════
    # SHEET 6: Unassigned Tracks
    # ═══════════════════════════════════════
    ws_un = wb.create_sheet('Unassigned Tracks')

    ws_un['A1'] = 'Tracks Not Assigned to Any Movement'
    ws_un['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_un['A2'] = ('After direct matching AND inference. Reasons explain why '
                   'each track could not be recovered automatically.')
    ws_un['A2'].font = Font(name='Arial', size=9, color='808080')

    # Get all non-stationary track IDs (also used by Tracks Between Gates sheet)
    all_tracks = query(
        "SELECT track_id, class_name, entry_edge, exit_edge, "
        "speed_mean_px, total_frames, first_frame, last_frame "
        "FROM tracks WHERE is_stationary=0 ORDER BY track_id")
    all_track_ids = set(t['track_id'] for t in all_tracks)

    unassigned = unresolved_tracks  # computed by movement_inference

    # Headers
    row = 4
    un_headers = ['Track ID', 'Class', 'Entry Edge', 'Exit Edge',
                  'Speed (px/f)', 'Frames', 'First Frame', 'Last Frame',
                  'Gates Crossed (with frames)', 'Reason']
    for c, h in enumerate(un_headers, 1):
        ws_un.cell(row=row, column=c, value=h)
        style_cell(ws_un.cell(row=row, column=c), font=header_font, fill=header_fill)

    ws_un.cell(row=3, column=1,
               value=f'Total unassigned: {len(unassigned)} out of {len(all_track_ids)} moving vehicles')
    ws_un.cell(row=3, column=1).font = Font(name='Arial', bold=True, size=10, color='C00000')

    row = 5
    for t in unassigned:
        ws_un.cell(row=row, column=1, value=t['track_id'])
        style_cell(ws_un.cell(row=row, column=1), align='left')
        ws_un.cell(row=row, column=2, value=t['class_name'])
        style_cell(ws_un.cell(row=row, column=2))
        ws_un.cell(row=row, column=3, value=t.get('entry_edge', ''))
        style_cell(ws_un.cell(row=row, column=3))
        ws_un.cell(row=row, column=4, value=t.get('exit_edge', ''))
        style_cell(ws_un.cell(row=row, column=4))
        ws_un.cell(row=row, column=5, value=t.get('speed', t.get('speed_mean_px', 0)))
        style_cell(ws_un.cell(row=row, column=5))
        ws_un.cell(row=row, column=6, value=t.get('total_frames', 0))
        style_cell(ws_un.cell(row=row, column=6))
        ws_un.cell(row=row, column=7, value=t.get('first_frame', 0))
        style_cell(ws_un.cell(row=row, column=7))
        ws_un.cell(row=row, column=8, value=t.get('last_frame', 0))
        style_cell(ws_un.cell(row=row, column=8))
        ws_un.cell(row=row, column=9, value=t.get('gates_crossed', ''))
        style_cell(ws_un.cell(row=row, column=9), align='left')
        ws_un.cell(row=row, column=10, value=t.get('reason', ''))
        style_cell(ws_un.cell(row=row, column=10), align='left')
        row += 1

    ws_un.column_dimensions['A'].width = 10
    ws_un.column_dimensions['B'].width = 8
    ws_un.column_dimensions['C'].width = 12
    ws_un.column_dimensions['D'].width = 12
    ws_un.column_dimensions['E'].width = 12
    ws_un.column_dimensions['F'].width = 8
    ws_un.column_dimensions['G'].width = 12
    ws_un.column_dimensions['H'].width = 12
    ws_un.column_dimensions['I'].width = 45
    ws_un.column_dimensions['J'].width = 55

    # ═══════════════════════════════════════
    # SHEET 7: Tracks Between Gates (missed by gate lines)
    # ═══════════════════════════════════════
    ws_between = wb.create_sheet('Tracks Between Gates')

    ws_between['A1'] = 'Tracks That Pass Between Gates But Missed Crossing'
    ws_between['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_between['A2'] = 'These tracks have trajectory points in the region between two gates but did not register a crossing on either gate line'
    ws_between['A2'].font = Font(name='Arial', size=9, color='808080')

    # For each movement pair, find tracks whose trajectory passes through
    # the bounding region of the two gates but didn't cross either
    import json as json_lib2

    # Get IDs that crossed each gate
    gate_crossing_ids = {}
    for g in all_gates:
        rows_gc = query("SELECT DISTINCT track_id FROM gate_crossings WHERE gate_id=?",
                        (g['gate_id'],))
        gate_crossing_ids[g['gate_id']] = set(r['track_id'] for r in rows_gc)

    col = 1
    for mov in movement_list:
        from_id = mov['from_id']
        to_id = mov['to_id']
        from_name = mov.get('from_name', str(from_id))
        to_name = mov.get('to_name', str(to_id))

        # Get gate coordinates
        gate_from = query("SELECT x1,y1,x2,y2 FROM gates WHERE gate_id=?", (from_id,), one=True)
        gate_to = query("SELECT x1,y1,x2,y2 FROM gates WHERE gate_id=?", (to_id,), one=True)
        if not gate_from or not gate_to:
            continue

        # Compute bounding box of the region between the two gates
        all_x = [gate_from['x1'], gate_from['x2'], gate_to['x1'], gate_to['x2']]
        all_y = [gate_from['y1'], gate_from['y2'], gate_to['y1'], gate_to['y2']]
        region_x_min = min(all_x)
        region_x_max = max(all_x)
        region_y_min = min(all_y)
        region_y_max = max(all_y)

        # Expand region slightly to catch near-misses at gate edges
        margin = 30
        region_x_min -= margin
        region_x_max += margin
        region_y_min -= margin
        region_y_max += margin

        # IDs already assigned to this movement (direct + inferred)
        crossed_from = gate_crossing_ids.get(from_id, set())
        crossed_to = gate_crossing_ids.get(to_id, set())
        assigned_to_movement = set(
            a['track_id'] for a in assignments.get((from_id, to_id), []))

        # Find tracks with trajectory points inside the region but NOT assigned
        between_tracks = []
        for t in all_tracks:
            tid = t['track_id']
            if tid in assigned_to_movement:
                continue

            # Check if trajectory passes through the region
            traj_row = query(
                "SELECT trajectory_json FROM tracks WHERE track_id=?", (tid,), one=True)
            if not traj_row or not traj_row.get('trajectory_json'):
                continue
            try:
                pts = json_lib2.loads(traj_row['trajectory_json'])
            except Exception:
                continue

            points_in_region = 0
            for p in pts:
                px, py = p[0], p[1]
                if region_x_min <= px <= region_x_max and region_y_min <= py <= region_y_max:
                    points_in_region += 1

            # Must have significant portion of trajectory in the region
            if points_in_region >= 3:
                crossed_which = []
                if tid in crossed_from:
                    crossed_which.append(from_name)
                if tid in crossed_to:
                    crossed_which.append(to_name)

                status = 'No gate crossed'
                if len(crossed_which) == 1:
                    status = f'Only crossed {crossed_which[0]}'

                between_tracks.append({
                    'track_id': tid,
                    'class_name': t['class_name'],
                    'entry_edge': t.get('entry_edge', ''),
                    'exit_edge': t.get('exit_edge', ''),
                    'speed': round(t.get('speed_mean_px', 0), 1),
                    'total_frames': t.get('total_frames', 0),
                    'points_in_region': points_in_region,
                    'total_points': len(pts),
                    'status': status,
                })

        # Write this movement's between-tracks
        movement_label = _mov_display(from_name, to_name).replace('→', '->')
        ws_between.cell(row=3, column=col,
                        value=f'{movement_label}: {len(between_tracks)} missed tracks')
        ws_between.cell(row=3, column=col).font = Font(
            name='Arial', bold=True, size=10, color='C00000')

        bw_headers = ['Track ID', 'Class', 'Entry', 'Exit', 'Speed', 'Frames',
                      'Pts in Region', 'Total Pts', 'Status']
        for c, h in enumerate(bw_headers, col):
            ws_between.cell(row=4, column=c, value=h)
            style_cell(ws_between.cell(row=4, column=c), font=header_font, fill=header_fill)

        for ri, bt in enumerate(between_tracks):
            r = 5 + ri
            ws_between.cell(row=r, column=col, value=bt['track_id'])
            style_cell(ws_between.cell(row=r, column=col), align='left')
            ws_between.cell(row=r, column=col + 1, value=bt['class_name'])
            style_cell(ws_between.cell(row=r, column=col + 1))
            ws_between.cell(row=r, column=col + 2, value=bt['entry_edge'])
            style_cell(ws_between.cell(row=r, column=col + 2))
            ws_between.cell(row=r, column=col + 3, value=bt['exit_edge'])
            style_cell(ws_between.cell(row=r, column=col + 3))
            ws_between.cell(row=r, column=col + 4, value=bt['speed'])
            style_cell(ws_between.cell(row=r, column=col + 4))
            ws_between.cell(row=r, column=col + 5, value=bt['total_frames'])
            style_cell(ws_between.cell(row=r, column=col + 5))
            ws_between.cell(row=r, column=col + 6, value=bt['points_in_region'])
            style_cell(ws_between.cell(row=r, column=col + 6))
            ws_between.cell(row=r, column=col + 7, value=bt['total_points'])
            style_cell(ws_between.cell(row=r, column=col + 7))
            ws_between.cell(row=r, column=col + 8, value=bt['status'])
            style_cell(ws_between.cell(row=r, column=col + 8), align='left')

        # Column widths
        widths = [10, 8, 8, 8, 8, 8, 12, 10, 30]
        for wi, w in enumerate(widths):
            c_letter = col + wi
            if c_letter <= 26:
                ws_between.column_dimensions[chr(64 + c_letter)].width = w

        col += 10  # gap between movements

    # ═══════════════════════════════════════
    # SHEET 8: Inference Report (QA trail)
    # ═══════════════════════════════════════
    _write_inference_report_sheet(wb, inference_report, style_cell,
                                  header_font, header_fill, data_font,
                                  total_font, total_fill)

    # ── Save ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'traffic_movements_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ═══════════════════════════════════════════════════════════
# TIME-BINNED MOVEMENT EXPORT
# ═══════════════════════════════════════════════════════════

@router.get("/export/time_range_info")
def time_range_info():
    """Return the video's available time range (for UI time pickers)."""
    from app.core.time_binned import get_time_range_info
    from app.core.database import get_conn
    return get_time_range_info(get_conn())


@router.get("/export/time_binned_excel")
def export_time_binned_excel(
    movements: str = '',
    bin_minutes: int = 15,
    range_start: str = None,
    range_end: str = None,
):
    """
    Export directional movement counts in time bins as a formatted Excel file.

    Query params:
        movements: JSON array [{from_id, from_name, to_id, to_name}, ...]
                   or empty/'auto' to generate all ordered leg-gate pairs
                   including U-turns
        bin_minutes: 1, 5, 15, 30, or 60 (default 15)
        range_start: start time filter — "08:00" or "08:00:00" or ISO datetime
        range_end: end time filter — "10:00" or ISO datetime
    """
    import json as json_lib
    if not HAS_OPENPYXL:
        raise HTTPException(500, "openpyxl not installed")

    from app.core.time_binned import compute_time_binned_movements
    from app.core.database import get_conn

    movement_list = _resolve_movements(movements)

    if bin_minutes not in (1, 5, 15, 30, 60):
        raise HTTPException(400, "bin_minutes must be 1, 5, 15, 30, or 60")

    conn = get_conn()
    result = compute_time_binned_movements(
        conn, movement_list,
        bin_minutes=bin_minutes,
        range_start=range_start,
        range_end=range_end,
    )

    time_bins = result['time_bins']
    movs = result['movements']
    class_names = result['class_names']
    grand_bins = result['grand_totals_per_bin']
    inference = result['inference']
    assignments = inference['assignments']
    unresolved_tracks = inference['unresolved']
    inference_report = inference['report']

    wb = Workbook()

    # ── Styles ──
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    sub_header_fill = PatternFill('solid', fgColor='D6E4F0')
    sub_header_font = Font(name='Arial', bold=True, size=10)
    data_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', bold=True, size=10, color='2F5496')
    total_fill = PatternFill('solid', fgColor='E2EFDA')
    time_font = Font(name='Arial', size=10, color='333333')
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'))

    def sc(cell, font=data_font, fill=None, align='center'):
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align)
        if fill:
            cell.fill = fill

    # ═══════════════════════════════════════
    # SHEET 1: Time-Binned Summary
    # ═══════════════════════════════════════
    ws = wb.active
    ws.title = 'Time-Binned Counts'

    ws['A1'] = 'Traffic Movement Counts — Time-Binned'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(name='Arial', size=9, color='808080')
    period_label = f'Period: {result["period_start"]} to {result["period_end"]}'
    ws['A3'] = (f'{period_label}  |  Bin: {bin_minutes} min  |  Excludes: Pedestrians  |  '
                f'Counts include inferred tracks (see Inference Report sheet)')
    ws['A3'].font = Font(name='Arial', size=9, color='808080')

    # Headers: Time Bin | Movement1 | Movement2 | ... | Grand Total
    row = 5
    ws.cell(row=row, column=1, value='Time Bin')
    sc(ws.cell(row=row, column=1), font=header_font, fill=header_fill)
    for mi, mov in enumerate(movs):
        ws.cell(row=row, column=2 + mi, value=mov['label'])
        sc(ws.cell(row=row, column=2 + mi), font=header_font, fill=header_fill)
    total_col = 2 + len(movs)
    ws.cell(row=row, column=total_col, value='Total')
    sc(ws.cell(row=row, column=total_col), font=header_font, fill=header_fill)

    # Data rows — one per time bin
    data_start_row = row + 1
    for bi, bin_label in enumerate(time_bins):
        r = data_start_row + bi
        ws.cell(row=r, column=1, value=bin_label)
        sc(ws.cell(row=r, column=1), font=time_font, align='left')

        for mi, mov in enumerate(movs):
            bin_total = sum(mov['bins'][bi].values())
            ws.cell(row=r, column=2 + mi, value=bin_total)
            sc(ws.cell(row=r, column=2 + mi))

        # Grand total for this bin = sum across all movements
        first_col_letter = chr(66)  # B
        last_col_letter = chr(66 + len(movs) - 1)
        ws.cell(row=r, column=total_col,
                value=f'=SUM({first_col_letter}{r}:{last_col_letter}{r})')
        sc(ws.cell(row=r, column=total_col), font=Font(name='Arial', bold=True, size=10))

    # Total row at bottom
    total_row = data_start_row + len(time_bins)
    ws.cell(row=total_row, column=1, value='TOTAL')
    sc(ws.cell(row=total_row, column=1), font=total_font, fill=total_fill, align='left')
    for col_idx in range(2, total_col + 1):
        col_letter = chr(64 + col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_letter}{data_start_row}:{col_letter}{total_row - 1})')
        sc(ws.cell(row=total_row, column=col_idx), font=total_font, fill=total_fill)

    # Column widths
    ws.column_dimensions['A'].width = 22
    for i in range(len(movs)):
        col_l = chr(66 + i)
        ws.column_dimensions[col_l].width = max(14, len(movs[i]['label']) + 2)
    ws.column_dimensions[chr(64 + total_col)].width = 10

    # Freeze top row
    ws.freeze_panes = 'A6'

    # ═══════════════════════════════════════
    # SHEET 2: Per-Movement Class Breakdown (time-binned)
    # ═══════════════════════════════════════
    for mov in movs:
        # Sheet name max 31 chars
        sheet_name = f"{mov['from_name']}→{mov['to_name']}"[:31]
        ws_m = wb.create_sheet(sheet_name)

        ws_m['A1'] = f"Movement: {mov['label']}"
        ws_m['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
        ws_m['A2'] = f'Bin: {bin_minutes} min  |  Excludes: Pedestrians'
        ws_m['A2'].font = Font(name='Arial', size=9, color='808080')

        # Headers: Time Bin | Class1 | Class2 | ... | Total
        r = 4
        headers_m = ['Time Bin'] + class_names + ['Total']
        for c, h in enumerate(headers_m, 1):
            ws_m.cell(row=r, column=c, value=h)
            sc(ws_m.cell(row=r, column=c), font=header_font, fill=header_fill)

        # Data
        m_data_start = r + 1
        for bi, bin_label in enumerate(time_bins):
            rr = m_data_start + bi
            ws_m.cell(row=rr, column=1, value=bin_label)
            sc(ws_m.cell(row=rr, column=1), font=time_font, align='left')

            for ci, cls in enumerate(class_names):
                count = mov['bins'][bi].get(cls, 0)
                ws_m.cell(row=rr, column=2 + ci, value=count)
                sc(ws_m.cell(row=rr, column=2 + ci))

            # Row total formula
            first_c = chr(66)  # B
            last_c = chr(66 + len(class_names) - 1)
            total_c = 2 + len(class_names)
            ws_m.cell(row=rr, column=total_c,
                      value=f'=SUM({first_c}{rr}:{last_c}{rr})')
            sc(ws_m.cell(row=rr, column=total_c),
               font=Font(name='Arial', bold=True, size=10))

        # Column totals
        m_total_row = m_data_start + len(time_bins)
        ws_m.cell(row=m_total_row, column=1, value='TOTAL')
        sc(ws_m.cell(row=m_total_row, column=1), font=total_font, fill=total_fill, align='left')
        for col_idx in range(2, 2 + len(class_names) + 1):
            col_letter = chr(64 + col_idx)
            ws_m.cell(row=m_total_row, column=col_idx,
                      value=f'=SUM({col_letter}{m_data_start}:{col_letter}{m_total_row - 1})')
            sc(ws_m.cell(row=m_total_row, column=col_idx), font=total_font, fill=total_fill)

        # Widths
        ws_m.column_dimensions['A'].width = 22
        for i in range(len(class_names)):
            ws_m.column_dimensions[chr(66 + i)].width = 10
        ws_m.column_dimensions[chr(66 + len(class_names))].width = 10
        ws_m.freeze_panes = 'A5'

    # ═══════════════════════════════════════
    # SHEET 3: Classification Grand Summary
    # ═══════════════════════════════════════
    ws_cls = wb.create_sheet('Classification Summary')

    ws_cls['A1'] = 'Vehicle Classification Summary (All Bins)'
    ws_cls['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_cls['A2'] = 'Direct = crossed both gates; Inferred = recovered single-crossing tracks'
    ws_cls['A2'].font = Font(name='Arial', size=9, color='808080')

    r = 4
    cls_headers = ['Movement'] + class_names + ['Direct', 'Inferred', 'Total']
    for c, h in enumerate(cls_headers, 1):
        ws_cls.cell(row=r, column=c, value=h)
        sc(ws_cls.cell(row=r, column=c), font=header_font, fill=header_fill)

    cls_data_start = r + 1
    n_cls = len(class_names)
    for mi, mov in enumerate(movs):
        rr = cls_data_start + mi
        ws_cls.cell(row=rr, column=1, value=mov['label'])
        sc(ws_cls.cell(row=rr, column=1), align='left')

        for ci, cls in enumerate(class_names):
            ws_cls.cell(row=rr, column=2 + ci, value=mov['total_per_class'].get(cls, 0))
            sc(ws_cls.cell(row=rr, column=2 + ci))

        # Direct / Inferred / Total
        ws_cls.cell(row=rr, column=2 + n_cls, value=mov['direct_total'])
        sc(ws_cls.cell(row=rr, column=2 + n_cls))
        ws_cls.cell(row=rr, column=3 + n_cls, value=mov['inferred_total'])
        sc(ws_cls.cell(row=rr, column=3 + n_cls),
           font=Font(name='Arial', size=10, color='B45309'))
        first_c = get_column_letter(2)
        last_c = get_column_letter(1 + n_cls)
        ws_cls.cell(row=rr, column=4 + n_cls,
                    value=f'=SUM({first_c}{rr}:{last_c}{rr})')
        sc(ws_cls.cell(row=rr, column=4 + n_cls),
           font=Font(name='Arial', bold=True, size=10))

    # Grand total row
    gt_row = cls_data_start + len(movs)
    ws_cls.cell(row=gt_row, column=1, value='GRAND TOTAL')
    sc(ws_cls.cell(row=gt_row, column=1), font=total_font, fill=total_fill, align='left')
    for col_idx in range(2, 5 + n_cls):
        col_letter = get_column_letter(col_idx)
        ws_cls.cell(row=gt_row, column=col_idx,
                    value=f'=SUM({col_letter}{cls_data_start}:{col_letter}{gt_row - 1})')
        sc(ws_cls.cell(row=gt_row, column=col_idx), font=total_font, fill=total_fill)

    ws_cls.column_dimensions['A'].width = 25
    for i in range(n_cls + 3):
        ws_cls.column_dimensions[get_column_letter(2 + i)].width = 10

    # ═══════════════════════════════════════
    # SHEET: Track IDs per Movement
    # ═══════════════════════════════════════
    ws_ids = wb.create_sheet('Track IDs by Movement')
    ws_ids['A1'] = 'Track IDs per Movement'
    ws_ids['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_ids['A2'] = 'Each column lists track IDs that crossed from one gate to another'
    ws_ids['A2'].font = Font(name='Arial', size=9, color='808080')

    import json as json_lib2
    id_col = 1
    all_assigned_ids = set()
    all_gates = query("SELECT * FROM gates")

    for mov in movement_list:
        from_id = mov['from_id']
        to_id = mov['to_id']
        from_name = mov.get('from_name', str(from_id))
        to_name = mov.get('to_name', str(to_id))

        id_rows = sorted(assignments.get((from_id, to_id), []),
                         key=lambda a: a['entry_frame'])
        n_direct = sum(1 for a in id_rows if not a['inferred'])
        n_inferred = len(id_rows) - n_direct

        hdr_lbl = _mov_display(from_name, to_name).replace('→', '->')
        ws_ids.cell(row=3, column=id_col,
                    value=f'Count: {len(id_rows)} ({n_direct} direct + {n_inferred} inferred)')
        ws_ids.cell(row=3, column=id_col).font = Font(name='Arial', bold=True, size=10, color='2F5496')

        for ci, h in enumerate(['Movement: ' + hdr_lbl, 'Class', 'Entry Frame',
                                'Exit Frame', 'Assigned By']):
            c = ws_ids.cell(row=4, column=id_col + ci, value=h)
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='2F5496')
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

        for ri, a in enumerate(id_rows):
            basis = 'Direct' if not a['inferred'] else f"Inferred: {a['basis']}"
            vals = [a['track_id'], a['class_name'], a['entry_frame'],
                    a['exit_frame'] if a['exit_frame'] is not None else '-', basis]
            for ci, val in enumerate(vals):
                c = ws_ids.cell(row=5 + ri, column=id_col + ci, value=val)
                c.font = (data_font if not a['inferred'] or ci < 4
                          else Font(name='Arial', size=10, color='B45309'))
                c.border = thin_border
            all_assigned_ids.add(a['track_id'])

        for wi, w in enumerate([22, 8, 12, 12, 22]):
            ws_ids.column_dimensions[get_column_letter(id_col + wi)].width = w

        id_col += 6

    # ═══════════════════════════════════════
    # SHEET: Unassigned Tracks
    # ═══════════════════════════════════════
    ws_un = wb.create_sheet('Unassigned Tracks')
    ws_un['A1'] = 'Tracks Not Assigned to Any Movement'
    ws_un['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_un['A2'] = ('After direct matching AND inference. Reasons explain why each '
                   'track could not be recovered automatically.')
    ws_un['A2'].font = Font(name='Arial', size=9, color='808080')

    all_tracks_list = query(
        "SELECT track_id, class_name, entry_edge, exit_edge, "
        "speed_mean_px, total_frames, first_frame, last_frame "
        "FROM tracks WHERE is_stationary=0 ORDER BY track_id")

    unassigned = unresolved_tracks  # computed by movement_inference

    ws_un.cell(row=3, column=1,
               value=f'Unassigned: {len(unassigned)} of {len(all_tracks_list)} moving vehicles')
    ws_un.cell(row=3, column=1).font = Font(name='Arial', bold=True, size=10, color='C00000')

    un_hdrs = ['Track ID', 'Class', 'Entry', 'Exit', 'Speed', 'Frames',
               'First Frame', 'Last Frame', 'Gates Crossed (with frames)', 'Reason']
    for ci, h in enumerate(un_hdrs):
        c = ws_un.cell(row=4, column=1 + ci, value=h)
        c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2F5496')
        c.border = thin_border

    for ri, t in enumerate(unassigned):
        vals = [t['track_id'], t['class_name'], t.get('entry_edge',''), t.get('exit_edge',''),
                t.get('speed', t.get('speed_mean_px', 0)), t.get('total_frames',0),
                t.get('first_frame',0), t.get('last_frame',0),
                t.get('gates_crossed',''), t.get('reason','')]
        for ci, val in enumerate(vals):
            c = ws_un.cell(row=5 + ri, column=1 + ci, value=val)
            c.font = data_font
            c.border = thin_border

    for ci, w in enumerate([10, 8, 10, 10, 8, 8, 12, 12, 45, 55]):
        ws_un.column_dimensions[chr(65 + ci)].width = w

    # ═══════════════════════════════════════
    # SHEET: Tracks Between Gates
    # ═══════════════════════════════════════
    ws_bw = wb.create_sheet('Tracks Between Gates')
    ws_bw['A1'] = 'Tracks Between Gates (missed crossing)'
    ws_bw['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_bw['A2'] = 'Trajectory passes through the region between gates but did not register a crossing'
    ws_bw['A2'].font = Font(name='Arial', size=9, color='808080')

    gate_crossing_ids = {}
    for g in all_gates:
        rows_gc = query("SELECT DISTINCT track_id FROM gate_crossings WHERE gate_id=?", (g['gate_id'],))
        gate_crossing_ids[g['gate_id']] = set(r['track_id'] for r in rows_gc)

    bw_col = 1
    for mov in movement_list:
        from_id = mov['from_id']
        to_id = mov['to_id']
        from_name = mov.get('from_name', str(from_id))
        to_name = mov.get('to_name', str(to_id))

        gf = query("SELECT x1,y1,x2,y2 FROM gates WHERE gate_id=?", (from_id,), one=True)
        gt_gate = query("SELECT x1,y1,x2,y2 FROM gates WHERE gate_id=?", (to_id,), one=True)
        if not gf or not gt_gate:
            continue

        ax = [gf['x1'], gf['x2'], gt_gate['x1'], gt_gate['x2']]
        ay = [gf['y1'], gf['y2'], gt_gate['y1'], gt_gate['y2']]
        margin = 30
        rx_min, rx_max = min(ax) - margin, max(ax) + margin
        ry_min, ry_max = min(ay) - margin, max(ay) + margin

        crossed_from = gate_crossing_ids.get(from_id, set())
        crossed_to = gate_crossing_ids.get(to_id, set())
        assigned_mov = set(a['track_id'] for a in assignments.get((from_id, to_id), []))

        between = []
        for t in all_tracks_list:
            tid = t['track_id']
            if tid in assigned_mov:
                continue
            traj_row = query("SELECT trajectory_json FROM tracks WHERE track_id=?", (tid,), one=True)
            if not traj_row or not traj_row.get('trajectory_json'):
                continue
            try:
                pts = json_lib2.loads(traj_row['trajectory_json'])
            except Exception:
                continue
            pts_in = sum(1 for p in pts if rx_min <= p[0] <= rx_max and ry_min <= p[1] <= ry_max)
            if pts_in >= 3:
                crossed_which = []
                if tid in crossed_from: crossed_which.append(from_name)
                if tid in crossed_to: crossed_which.append(to_name)
                status = 'No gate crossed'
                if len(crossed_which) == 1:
                    status = f'Only crossed {crossed_which[0]}'
                between.append({
                    'track_id': tid, 'class': t['class_name'],
                    'entry': t.get('entry_edge',''), 'exit': t.get('exit_edge',''),
                    'speed': round(t.get('speed_mean_px',0), 1),
                    'frames': t.get('total_frames',0),
                    'pts_in': pts_in, 'total_pts': len(pts), 'status': status})

        lbl = _mov_display(from_name, to_name).replace('→', '->')
        ws_bw.cell(row=3, column=bw_col, value=f'{lbl}: {len(between)} missed')
        ws_bw.cell(row=3, column=bw_col).font = Font(name='Arial', bold=True, size=10, color='C00000')

        bw_hdrs = ['Track ID', 'Class', 'Entry', 'Exit', 'Speed', 'Frames', 'Pts In Region', 'Total Pts', 'Status']
        for ci, h in enumerate(bw_hdrs):
            c = ws_bw.cell(row=4, column=bw_col + ci, value=h)
            c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='2F5496')
            c.border = thin_border

        for ri, bt in enumerate(between):
            vals = [bt['track_id'], bt['class'], bt['entry'], bt['exit'],
                    bt['speed'], bt['frames'], bt['pts_in'], bt['total_pts'], bt['status']]
            for ci, val in enumerate(vals):
                c = ws_bw.cell(row=5 + ri, column=bw_col + ci, value=val)
                c.font = data_font
                c.border = thin_border

        bw_col += 10

    # ═══════════════════════════════════════
    # SHEET: Inference Report (QA trail)
    # ═══════════════════════════════════════
    _write_inference_report_sheet(wb, inference_report, sc,
                                  header_font, header_fill, data_font,
                                  total_font, total_fill)

    # ── Save ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'traffic_time_binned_{bin_minutes}min_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
