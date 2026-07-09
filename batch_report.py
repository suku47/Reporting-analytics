"""
batch_report.py — One-command TMC extraction for a whole location.

Workflow this replaces: opening every .traf in the viewer, drawing the same
gates, exporting Excel per file, then merging.

New workflow:
  1. Draw gates ONCE on the first file in the viewer (as today).
  2. Run:
     python batch_report.py --gates-from Results/traf/VID_075214.traf \
                            --traf-dir Results/traf \
                            --bin 15 \
                            --period 08:00-10:00 --period 15:00-17:00 \
                            --out Results/reports --site "Chester Site 4"

What it does, per .traf in --traf-dir:
  - copies the reference gates in (wiping any old gates first)
  - recomputes ALL crossings headlessly (same gate_engine the viewer uses)
  - computes time-binned movement counts (same time_binned engine)
Then merges everything across files (bins summed by clock label) and writes
ONE merged workbook per survey period (or one for the full recording when no
--period is given), styled like the viewer's own TMC export, plus a
'Files' QA sheet showing per-file contributions.

The .traf files keep their gates + crossings, so any file can still be
opened in the viewer afterwards for spot-checking.
"""

import argparse
import os
import re
import sqlite3
import sys
import glob
from datetime import datetime, timedelta

# Run from the TrafficAnalyticsViewer folder (same place as run.py)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.core.gate_engine import recompute_all_gates
from app.core.time_binned import compute_time_binned_movements
from app.core.movement_inference import auto_generate_movements

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required.  pip install openpyxl")
    sys.exit(1)


# ══════════════════════════════════════════════════════════════════════
#  Auto movement numbering (client compass convention)
# ══════════════════════════════════════════════════════════════════════

# Clockwise compass ring
_RING = ['N', 'E', 'S', 'W']
# Client convention (confirmed): approaches processed clockwise in bound
# order SB, WB, NB, EB — i.e. ARMS in order N, E, S, W; letters follow the
# same order (A=N, B=E, C=S, D=W). Turn order for left-hand (UK) traffic:
# Left, Through, Right, U-Turn. Absent arms are skipped; numbering is
# continuous.
DEFAULT_ARM_ORDER = ['N', 'E', 'S', 'W']
DEFAULT_ARM_LETTERS = {'N': 'A', 'E': 'B', 'S': 'C', 'W': 'D'}
DEFAULT_TURN_ORDER = ['L', 'T', 'R', 'U']


def canon_compass(name):
    """'North'/'north'/'N' → 'N'; returns None if not a compass name."""
    return {'n': 'N', 'north': 'N', 's': 'S', 'south': 'S',
            'e': 'E', 'east': 'E', 'w': 'W', 'west': 'W'}.get(
                str(name).strip().lower())


def turn_exits(from_arm):
    """Exit arm for each turn from an approach arm (geometry, not drive side):
    a vehicle entering from arm X heads towards opposite(X);
    Left exit = clockwise-next(X), Through = opposite, Right = ccw-next,
    U = X itself."""
    i = _RING.index(from_arm)
    return {'L': _RING[(i + 1) % 4], 'T': _RING[(i + 2) % 4],
            'R': _RING[(i + 3) % 4], 'U': from_arm}


def generate_client_movements(gate_names, site_number,
                              arm_order=None, turn_order=None,
                              arm_letters=None, arm_modes=None):
    """Given compass-named gates present at a site, produce the client
    movement-number mapping, e.g. {'South→West': '59.4', ...}.

    gate_names: iterable of gate names as drawn in the viewer
                (must canonicalise to N/S/E/W)
    site_number: the client's site prefix (e.g. 59 → 59.1, 59.2, ...)
    arm_modes: optional dict {arm: 'two'|'in'|'out'} (canonical or full
               compass names accepted). Default: every arm 'two' (two-way).
               'in'  = one-way towards the junction: arm is an approach but
                       NO movement may exit via it (incl. its own U-turn)
               'out' = one-way away from the junction: no approach from it,
                       but other approaches may exit into it

    Rule (confirmed against three client examples): movement X→Y exists iff
    X allows flow into the junction and Y allows flow out; approaches in
    SB, WB, NB, EB order; turns in L, T, R, U order; numbering continuous
    across all skips.

    Returns (mapping dict keyed 'From→To' using the ORIGINAL gate names,
             table rows for printing/verification).
    """
    arm_order = arm_order or DEFAULT_ARM_ORDER
    turn_order = turn_order or DEFAULT_TURN_ORDER
    arm_letters = arm_letters or DEFAULT_ARM_LETTERS

    # Map canonical arm → original gate name (preserve user's naming)
    arm_to_gate = {}
    bad = []
    for gname in gate_names:
        arm = canon_compass(gname)
        if arm is None:
            bad.append(str(gname))
        elif arm in arm_to_gate:
            bad.append(f"{gname} (duplicate of {arm_to_gate[arm]})")
        else:
            arm_to_gate[arm] = str(gname)
    if bad:
        raise ValueError(
            "Auto-movements needs gates named by compass "
            "(North/South/East/West or N/S/E/W). Problem gate(s): "
            + ", ".join(bad))

    # Canonicalise arm modes; default two-way
    modes = {arm: 'two' for arm in arm_to_gate}
    for k, v in (arm_modes or {}).items():
        arm = canon_compass(k)
        if arm is None or arm not in arm_to_gate:
            raise ValueError(f"--arm-mode: unknown arm '{k}' "
                             f"(no gate with that compass name)")
        v = str(v).strip().lower()
        if v not in ('two', 'in', 'out'):
            raise ValueError(f"--arm-mode for '{k}' must be two/in/out, got '{v}'")
        modes[arm] = v

    def can_approach(arm):
        return modes[arm] in ('two', 'in')

    def can_exit(arm):
        return modes[arm] in ('two', 'out')

    mapping = {}
    rows = []
    k = 0
    for arm in arm_order:
        if arm not in arm_to_gate or not can_approach(arm):
            continue
        exits = turn_exits(arm)
        for turn in turn_order:
            exit_arm = exits[turn]
            if exit_arm not in arm_to_gate or not can_exit(exit_arm):
                continue  # absent/one-way-blocked → skipped, numbering continues
            k += 1
            num = f"{site_number}.{k}"
            key = f"{arm_to_gate[arm]}→{arm_to_gate[exit_arm]}"
            mapping[key] = num
            rows.append((num,
                         f"{arm_letters[arm]}→{arm_letters[exit_arm]}",
                         key,
                         {'L': 'Left', 'T': 'Through',
                          'R': 'Right', 'U': 'U-turn'}[turn]))
    return mapping, rows



GATES_SCHEMA = """
CREATE TABLE IF NOT EXISTS gates (
    gate_id     INTEGER PRIMARY KEY AUTOINCREMENT,
    name        TEXT,
    x1          REAL, y1 REAL, x2 REAL, y2 REAL,
    direction   TEXT,
    approach    TEXT,
    created_at  TEXT,
    arm_mode    TEXT DEFAULT 'two'
)"""

CROSSINGS_SCHEMA = """
CREATE TABLE IF NOT EXISTS gate_crossings (
    crossing_id INTEGER PRIMARY KEY AUTOINCREMENT,
    gate_id     INTEGER,
    track_id    INTEGER,
    frame       INTEGER,
    timestamp   TEXT,
    direction   TEXT,
    speed_px    REAL,
    class_id    INTEGER,
    class_name  TEXT
)"""


def read_reference_gates(ref_path):
    """Read gates (incl. arm_mode) + frame size from the reference .traf."""
    conn = sqlite3.connect(ref_path)
    try:
        gates = conn.execute(
            "SELECT name, x1, y1, x2, y2, direction, approach, "
            "COALESCE(arm_mode, 'two') FROM gates ORDER BY gate_id").fetchall()
    except sqlite3.OperationalError:
        # Older traf without arm_mode column
        gates = [(*g, 'two') for g in conn.execute(
            "SELECT name, x1, y1, x2, y2, direction, approach "
            "FROM gates ORDER BY gate_id").fetchall()]
    meta = {r[0]: r[1] for r in conn.execute("SELECT key, value FROM scene")}
    conn.close()
    if not gates:
        sys.exit(f"ERROR: no gates found in reference file {ref_path}.\n"
                 f"Open it in the viewer and draw the gates first.")
    frame_w = int(meta.get('frame_width', 0))
    frame_h = int(meta.get('frame_height', 0))
    return gates, frame_w, frame_h


def apply_gates(conn, gates):
    """Wipe existing gates/crossings and insert the reference gates in order."""
    conn.execute(GATES_SCHEMA)
    conn.execute(CROSSINGS_SCHEMA)
    try:
        conn.execute("ALTER TABLE gates ADD COLUMN arm_mode TEXT DEFAULT 'two'")
    except sqlite3.OperationalError:
        pass  # column already exists (or table just created with it)
    conn.execute("DELETE FROM gate_crossings")
    conn.execute("DELETE FROM gates")
    now = datetime.now().isoformat()
    for g in gates:
        conn.execute(
            "INSERT INTO gates (name, x1, y1, x2, y2, direction, approach, "
            "created_at, arm_mode) VALUES (?,?,?,?,?,?,?,?,?)",
            (*g[:7], now, g[7]))
    conn.commit()


# ══════════════════════════════════════════════════════════════════════
#  Merging per-file results
# ══════════════════════════════════════════════════════════════════════

def _bin_sort_key(label):
    """Sort 'HH:MM - HH:MM' labels chronologically."""
    m = re.match(r'(\d{1,2}):(\d{2})', label)
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)


def merge_results(per_file_results, class_order_hint=None):
    """Merge a list of compute_time_binned_movements() results into one
    result dict of the same shape (bins summed by clock label)."""
    if not per_file_results:
        return None

    # Union of classes (stable order: hint first, then discovery order)
    class_names = list(class_order_hint or [])
    for res in per_file_results:
        for c in res['class_names']:
            if c not in class_names:
                class_names.append(c)

    # Union of bin labels, chronological
    bin_labels = []
    for res in per_file_results:
        for b in res['time_bins']:
            if b not in bin_labels:
                bin_labels.append(b)
    bin_labels.sort(key=_bin_sort_key)
    bin_index = {b: i for i, b in enumerate(bin_labels)}

    # Movements keyed by (from_name, to_name), order of first appearance
    merged_movs = {}
    mov_order = []
    for res in per_file_results:
        for mov in res['movements']:
            key = (mov['from_name'], mov['to_name'])
            if key not in merged_movs:
                mov_order.append(key)
                merged_movs[key] = {
                    'from_id': mov['from_id'], 'to_id': mov['to_id'],
                    'from_name': mov['from_name'], 'to_name': mov['to_name'],
                    'label': mov['label'],
                    'bins': [{c: 0 for c in class_names} for _ in bin_labels],
                    'total_per_class': {c: 0 for c in class_names},
                    'direct_per_class': {c: 0 for c in class_names},
                    'inferred_per_class': {c: 0 for c in class_names},
                    'direct_total': 0, 'inferred_total': 0, 'grand_total': 0,
                }
            tgt = merged_movs[key]
            for bi, src_bin in enumerate(mov['bins']):
                gi = bin_index[res['time_bins'][bi]]
                for c, n in src_bin.items():
                    if n:
                        tgt['bins'][gi][c] = tgt['bins'][gi].get(c, 0) + n
            for field in ('total_per_class', 'direct_per_class', 'inferred_per_class'):
                for c, n in mov[field].items():
                    if n:
                        tgt[field][c] = tgt[field].get(c, 0) + n
            tgt['direct_total'] += mov['direct_total']
            tgt['inferred_total'] += mov['inferred_total']
            tgt['grand_total'] += mov['grand_total']

    movements = [merged_movs[k] for k in mov_order]

    grand_bins = [{c: 0 for c in class_names} for _ in bin_labels]
    for mov in movements:
        for bi, b in enumerate(mov['bins']):
            for c in class_names:
                grand_bins[bi][c] += b.get(c, 0)

    period_starts = [r['period_start'] for r in per_file_results]
    period_ends = [r['period_end'] for r in per_file_results]

    return {
        'time_bins': bin_labels,
        'bin_minutes': per_file_results[0]['bin_minutes'],
        'movements': movements,
        'class_names': class_names,
        'grand_totals_per_bin': grand_bins,
        'period_start': min(period_starts),
        'period_end': max(period_ends),
    }


# ══════════════════════════════════════════════════════════════════════
#  Merged workbook (styled like the viewer's TMC export)
# ══════════════════════════════════════════════════════════════════════

def build_merged_workbook(result, site_label, period_label, file_stats):
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    data_font = Font(name='Arial', size=10)
    total_font = Font(name='Arial', bold=True, size=10, color='2F5496')
    total_fill = PatternFill('solid', fgColor='E2EFDA')
    time_font = Font(name='Arial', size=10, color='333333')
    thin_border = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'))

    def sc(cell, font=data_font, fill=None, align='center'):
        cell.font = font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal=align)
        if fill:
            cell.fill = fill

    time_bins = result['time_bins']
    movs = result['movements']
    class_names = result['class_names']
    bin_minutes = result['bin_minutes']

    wb = Workbook()

    # ── SHEET 1: Time-Binned Counts (all movements) ──
    ws = wb.active
    ws.title = 'Time-Binned Counts'
    ws['A1'] = f'{site_label} — Merged Traffic Movement Counts'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws['A2'] = (f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}  |  '
                f'Merged from {len(file_stats)} file(s)')
    ws['A2'].font = Font(name='Arial', size=9, color='808080')
    ws['A3'] = (f'Period: {period_label}  |  Bin: {bin_minutes} min  |  '
                f'Excludes: Pedestrians  |  Counts include inferred tracks')
    ws['A3'].font = Font(name='Arial', size=9, color='808080')

    row = 5
    ws.cell(row=row, column=1, value='Time Bin')
    sc(ws.cell(row=row, column=1), font=header_font, fill=header_fill)
    for mi, mov in enumerate(movs):
        ws.cell(row=row, column=2 + mi, value=mov['label'])
        sc(ws.cell(row=row, column=2 + mi), font=header_font, fill=header_fill)
    total_col = 2 + len(movs)
    ws.cell(row=row, column=total_col, value='Total')
    sc(ws.cell(row=row, column=total_col), font=header_font, fill=header_fill)

    data_start_row = row + 1
    for bi, bin_label in enumerate(time_bins):
        r = data_start_row + bi
        ws.cell(row=r, column=1, value=bin_label)
        sc(ws.cell(row=r, column=1), font=time_font, align='left')
        for mi, mov in enumerate(movs):
            ws.cell(row=r, column=2 + mi, value=sum(mov['bins'][bi].values()))
            sc(ws.cell(row=r, column=2 + mi))
        first_l = get_column_letter(2)
        last_l = get_column_letter(1 + len(movs))
        ws.cell(row=r, column=total_col, value=f'=SUM({first_l}{r}:{last_l}{r})')
        sc(ws.cell(row=r, column=total_col), font=Font(name='Arial', bold=True, size=10))

    total_row = data_start_row + len(time_bins)
    ws.cell(row=total_row, column=1, value='TOTAL')
    sc(ws.cell(row=total_row, column=1), font=total_font, fill=total_fill, align='left')
    for col_idx in range(2, total_col + 1):
        col_l = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx,
                value=f'=SUM({col_l}{data_start_row}:{col_l}{total_row - 1})')
        sc(ws.cell(row=total_row, column=col_idx), font=total_font, fill=total_fill)

    ws.column_dimensions['A'].width = 22
    for i, mov in enumerate(movs):
        ws.column_dimensions[get_column_letter(2 + i)].width = max(14, len(mov['label']) + 2)
    ws.column_dimensions[get_column_letter(total_col)].width = 10
    ws.freeze_panes = 'A6'

    # ── SHEETS: per-movement class breakdown ──
    for mov in movs:
        sheet_name = f"{mov['from_name']}→{mov['to_name']}"[:31]
        ws_m = wb.create_sheet(sheet_name)
        ws_m['A1'] = f"Movement: {mov['label']}"
        ws_m['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
        ws_m['A2'] = f'Bin: {bin_minutes} min  |  Excludes: Pedestrians  |  Merged across files'
        ws_m['A2'].font = Font(name='Arial', size=9, color='808080')

        r = 4
        for c, h in enumerate(['Time Bin'] + class_names + ['Total'], 1):
            ws_m.cell(row=r, column=c, value=h)
            sc(ws_m.cell(row=r, column=c), font=header_font, fill=header_fill)

        m_start = r + 1
        for bi, bin_label in enumerate(time_bins):
            rr = m_start + bi
            ws_m.cell(row=rr, column=1, value=bin_label)
            sc(ws_m.cell(row=rr, column=1), font=time_font, align='left')
            for ci, cls in enumerate(class_names):
                ws_m.cell(row=rr, column=2 + ci, value=mov['bins'][bi].get(cls, 0))
                sc(ws_m.cell(row=rr, column=2 + ci))
            first_c = get_column_letter(2)
            last_c = get_column_letter(1 + len(class_names))
            ws_m.cell(row=rr, column=2 + len(class_names),
                      value=f'=SUM({first_c}{rr}:{last_c}{rr})')
            sc(ws_m.cell(row=rr, column=2 + len(class_names)),
               font=Font(name='Arial', bold=True, size=10))

        m_total = m_start + len(time_bins)
        ws_m.cell(row=m_total, column=1, value='TOTAL')
        sc(ws_m.cell(row=m_total, column=1), font=total_font, fill=total_fill, align='left')
        for col_idx in range(2, 3 + len(class_names)):
            col_l = get_column_letter(col_idx)
            ws_m.cell(row=m_total, column=col_idx,
                      value=f'=SUM({col_l}{m_start}:{col_l}{m_total - 1})')
            sc(ws_m.cell(row=m_total, column=col_idx), font=total_font, fill=total_fill)

        ws_m.column_dimensions['A'].width = 22
        for i in range(len(class_names) + 1):
            ws_m.column_dimensions[get_column_letter(2 + i)].width = 10
        ws_m.freeze_panes = 'A5'

    # ── SHEET: Classification Summary ──
    ws_cls = wb.create_sheet('Classification Summary')
    ws_cls['A1'] = 'Vehicle Classification Summary (All Bins, Merged)'
    ws_cls['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_cls['A2'] = 'Direct = crossed both gates; Inferred = recovered single-crossing tracks'
    ws_cls['A2'].font = Font(name='Arial', size=9, color='808080')

    r = 4
    n_cls = len(class_names)
    for c, h in enumerate(['Movement'] + class_names + ['Direct', 'Inferred', 'Total'], 1):
        ws_cls.cell(row=r, column=c, value=h)
        sc(ws_cls.cell(row=r, column=c), font=header_font, fill=header_fill)

    cls_start = r + 1
    for mi, mov in enumerate(movs):
        rr = cls_start + mi
        ws_cls.cell(row=rr, column=1, value=mov['label'])
        sc(ws_cls.cell(row=rr, column=1), align='left')
        for ci, cls in enumerate(class_names):
            ws_cls.cell(row=rr, column=2 + ci, value=mov['total_per_class'].get(cls, 0))
            sc(ws_cls.cell(row=rr, column=2 + ci))
        ws_cls.cell(row=rr, column=2 + n_cls, value=mov['direct_total'])
        sc(ws_cls.cell(row=rr, column=2 + n_cls))
        ws_cls.cell(row=rr, column=3 + n_cls, value=mov['inferred_total'])
        sc(ws_cls.cell(row=rr, column=3 + n_cls),
           font=Font(name='Arial', size=10, color='B45309'))
        first_c = get_column_letter(2)
        last_c = get_column_letter(1 + n_cls)
        ws_cls.cell(row=rr, column=4 + n_cls, value=f'=SUM({first_c}{rr}:{last_c}{rr})')
        sc(ws_cls.cell(row=rr, column=4 + n_cls), font=Font(name='Arial', bold=True, size=10))

    gt_row = cls_start + len(movs)
    ws_cls.cell(row=gt_row, column=1, value='GRAND TOTAL')
    sc(ws_cls.cell(row=gt_row, column=1), font=total_font, fill=total_fill, align='left')
    for col_idx in range(2, 5 + n_cls):
        col_l = get_column_letter(col_idx)
        ws_cls.cell(row=gt_row, column=col_idx,
                    value=f'=SUM({col_l}{cls_start}:{col_l}{gt_row - 1})')
        sc(ws_cls.cell(row=gt_row, column=col_idx), font=total_font, fill=total_fill)

    ws_cls.column_dimensions['A'].width = 25
    for i in range(n_cls + 3):
        ws_cls.column_dimensions[get_column_letter(2 + i)].width = 10

    # ── SHEET: Files (QA trail) ──
    ws_f = wb.create_sheet('Files')
    ws_f['A1'] = 'Per-File Contributions (QA)'
    ws_f['A1'].font = Font(name='Arial', bold=True, size=12, color='2F5496')
    ws_f['A2'] = ('Counts contributed by each source file to this period. '
                  'Open any file in the viewer to spot-check its gates/crossings.')
    ws_f['A2'].font = Font(name='Arial', size=9, color='808080')

    r = 4
    f_hdrs = ['File', 'Video Start', 'Video End', 'Moving Tracks',
              'Crossings', 'Counted (Direct)', 'Counted (Inferred)', 'Counted Total']
    for c, h in enumerate(f_hdrs, 1):
        ws_f.cell(row=r, column=c, value=h)
        sc(ws_f.cell(row=r, column=c), font=header_font, fill=header_fill)

    f_start = r + 1
    for fi, st in enumerate(file_stats):
        rr = f_start + fi
        vals = [st['name'], st['video_start'], st['video_end'], st['tracks'],
                st['crossings'], st['direct'], st['inferred'],
                st['direct'] + st['inferred']]
        for ci, v in enumerate(vals, 1):
            ws_f.cell(row=rr, column=ci, value=v)
            sc(ws_f.cell(row=rr, column=ci), align='left' if ci <= 3 else 'center')

    t_row = f_start + len(file_stats)
    ws_f.cell(row=t_row, column=1, value='TOTAL')
    sc(ws_f.cell(row=t_row, column=1), font=total_font, fill=total_fill, align='left')
    for col_idx in range(4, 9):
        col_l = get_column_letter(col_idx)
        ws_f.cell(row=t_row, column=col_idx,
                  value=f'=SUM({col_l}{f_start}:{col_l}{t_row - 1})')
        sc(ws_f.cell(row=t_row, column=col_idx), font=total_font, fill=total_fill)

    ws_f.column_dimensions['A'].width = 32
    for col_l, w in (('B', 20), ('C', 20), ('D', 14), ('E', 12), ('F', 16), ('G', 16), ('H', 14)):
        ws_f.column_dimensions[col_l].width = w

    return wb


# ══════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════

def parse_period(s):
    """'08:00-10:00' → ('08:00', '10:00')"""
    m = re.match(r'^(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})$', s.strip())
    if not m:
        raise argparse.ArgumentTypeError(
            f"Invalid period '{s}' — use HH:MM-HH:MM, e.g. 08:00-10:00")
    return m.group(1), m.group(2)


SITE_CONFIG_KEYS = """
Site config YAML keys (all CLI options have a YAML equivalent; CLI overrides YAML):
  site: Chester Site 4              # report title label
  gates_from: path/to/reference.traf
  traf_dir: path/to/traf/folder
  out: path/to/reports              # output folder
  bin: 15                           # 1/5/15/30/60
  periods: ["08:00-10:00", "15:00-17:00"]   # omit for full recording
  site_number: 59                   # enables auto movement numbering
  arm_modes: {North: in}            # optional one-way overrides (else from gates)
  client_template: path/to/client.xlsm      # optional
  classes: {Car: "Cars (LMV)"}      # class map (inline, preferred)
  mapping: path/to/mapping.json     # OR external mapping file
  movements: {"C->D": "59.4"}       # manual movement map (if no site_number)
  schedule: path/to/Video_schedule.xlsx     # optional, enables QA reconciliation
"""


def load_site_config(path):
    try:
        import yaml
    except ImportError:
        raise RuntimeError("PyYAML required for --config: pip install pyyaml")
    with open(path, encoding='utf-8') as f:
        cfg = yaml.safe_load(f) or {}
    if not isinstance(cfg, dict):
        raise RuntimeError(f"Config {path} must be a YAML mapping")
    return cfg


# ══════════════════════════════════════════════════════════════════════
#  QA checks
# ══════════════════════════════════════════════════════════════════════

def _qa_gate_direction_check(conn, warnings, fname):
    """Compare each gate's declared arm_mode with observed crossing
    directions: a two-way arm should see both directions; a one-way arm
    should be heavily one-sided."""
    try:
        gates = conn.execute(
            "SELECT gate_id, name, COALESCE(arm_mode,'two') FROM gates").fetchall()
    except Exception:
        return
    for gid, gname, mode in gates:
        rows = conn.execute(
            "SELECT direction, COUNT(*) FROM gate_crossings WHERE gate_id=? "
            "GROUP BY direction", (gid,)).fetchall()
        total = sum(r[1] for r in rows)
        if total < 20:
            continue  # too little data to judge
        top = max(r[1] for r in rows)
        one_sidedness = top / total
        if mode == 'two' and one_sidedness > 0.97:
            warnings.append(
                f"{fname}: gate '{gname}' is set two-way but {one_sidedness:.0%} "
                f"of {total} crossings are one direction — is this arm one-way?")
        elif mode in ('in', 'out') and one_sidedness < 0.80:
            warnings.append(
                f"{fname}: gate '{gname}' is set one-way ({mode}) but crossings "
                f"go both ways ({one_sidedness:.0%} majority of {total}) — "
                f"check the arm mode.")


def _qa_schedule_reconciliation(schedule_path, traf_files, warnings):
    """Every scheduled video should have a .traf; flag any that don't."""
    try:
        import pandas as pd
        df = pd.read_excel(schedule_path)
        scheduled = [os.path.splitext(os.path.basename(str(p)))[0]
                     for p in df['video_path']]
    except Exception as e:
        warnings.append(f"QA: could not read schedule for reconciliation ({e})")
        return
    traf_stems = {os.path.splitext(os.path.basename(p))[0] for p in traf_files}
    missing = [s for s in scheduled if s not in traf_stems]
    if missing:
        warnings.append(
            f"SCHEDULE MISMATCH: {len(missing)} scheduled video(s) have NO .traf "
            f"in the folder (processing failed or still running?): "
            + ", ".join(missing[:10]) + (" …" if len(missing) > 10 else ""))
    extra = sorted(traf_stems - set(scheduled))
    if extra:
        warnings.append(
            f"QA note: {len(extra)} .traf file(s) not in the schedule "
            f"(from another run?): " + ", ".join(extra[:10]))


def _qa_period_coverage(periods, qa_files, warnings):
    """Warn if the recordings don't cover the requested survey periods."""
    spans = []
    for f in qa_files:
        try:
            s = datetime.fromisoformat(f['video_start'])
            e = datetime.fromisoformat(f['video_end'])
            spans.append((s, e))
        except (ValueError, TypeError, KeyError):
            continue
    if not spans:
        return
    for p_start, p_end in periods:
        if not p_start:
            continue
        ref_date = spans[0][0].date()
        ps = datetime.combine(ref_date, datetime.strptime(p_start, '%H:%M').time())
        pe = datetime.combine(ref_date, datetime.strptime(p_end, '%H:%M').time())
        covered = 0.0
        for s, e in spans:
            lo, hi = max(s, ps), min(e, pe)
            if hi > lo:
                covered += (hi - lo).total_seconds()
        needed = (pe - ps).total_seconds()
        if needed > 0 and covered < needed * 0.999:
            gap_min = (needed - covered) / 60
            warnings.append(
                f"COVERAGE GAP: period {p_start}-{p_end} has {gap_min:.0f} min "
                f"NOT covered by any recording — those bins will read low/zero.")


def run_qa(qa_files, warnings):
    """Per-file outlier checks. Adds flags into qa_files rows and warnings."""
    import statistics
    cpms = [f['cpm'] for f in qa_files if f['cpm'] is not None and f['cpm'] > 0]
    median_cpm = statistics.median(cpms) if cpms else 0
    for f in qa_files:
        flags = []
        if f['crossings'] == 0:
            flags.append('NO CROSSINGS')
        elif median_cpm and f['cpm'] is not None and f['cpm'] < 0.3 * median_cpm:
            flags.append(f"LOW ({f['cpm']:.1f}/min vs median {median_cpm:.1f})")
        if f.get('skipped'):
            flags.append('SKIPPED: ' + f['skipped'])
        f['flags'] = flags
        if flags:
            warnings.append(f"{f['name']}: " + "; ".join(flags))
    return median_cpm


def add_qa_sheet(wb, qa_files, warnings, median_cpm):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    warn_font = Font(name='Arial', size=10, color='B42318')
    ok_font = Font(name='Arial', size=10, color='1F7A33')
    data_font = Font(name='Arial', size=10)
    thin = Side(style='thin', color='B4C6E7')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet('QA')
    ws['A1'] = 'Automated QA Checks'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    status = ('PASS — no issues found' if not warnings
              else f'{len(warnings)} WARNING(S) — review before delivery')
    ws['A2'] = status
    ws['A2'].font = ok_font if not warnings else Font(
        name='Arial', bold=True, size=11, color='B42318')

    r = 4
    for w in warnings:
        ws.cell(row=r, column=1, value='⚠ ' + w).font = warn_font
        r += 1
    r += 1

    hdrs = ['File', 'Duration (min)', 'Moving Tracks', 'Crossings',
            'Crossings/min', 'Flags']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    r += 1
    for f in qa_files:
        vals = [f['name'], round(f['dur_min'], 1) if f['dur_min'] else '—',
                f['tracks'], f['crossings'],
                round(f['cpm'], 2) if f['cpm'] is not None else '—',
                '; '.join(f['flags']) if f['flags'] else 'OK']
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = warn_font if f['flags'] and c == len(vals) else data_font
            cell.border = border
        r += 1
    ws.cell(row=r, column=1, value=f'Site median crossings/min: '
            f'{median_cpm:.2f}').font = data_font
    ws.column_dimensions['A'].width = 32
    for i, w in enumerate([16, 14, 12, 14, 46], 2):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════
#  Site-level trajectory plots (merged across all files, per period)
# ══════════════════════════════════════════════════════════════════════

def generate_site_trajectory_plots(cfg, traf_files, periods, out_dir,
                                   site_label, log=print):
    """One trajectory PNG per period, merging every file's tracks onto a
    single background frame — the trajectory-plot equivalent of the merged
    TMC workbook. Returns list of written image paths."""
    from app.core.trajectory_render import (
        read_trajectories, read_scene_meta, grab_background_frame,
        despike_trajectories, smooth_trajectories, trim_departing_starts,
        filter_noisy_in_crowded_approaches, draw_trajectories,
        draw_legend, draw_job_info)
    import numpy as np
    import cv2
    import random

    # ── Background frame: explicit override, else first schedule video ──
    bg_video = cfg.get('background_video')
    if not bg_video and cfg.get('schedule'):
        try:
            import pandas as pd
            df = pd.read_excel(cfg['schedule'])
            for p in df['video_path']:
                if os.path.exists(str(p)):
                    bg_video = str(p)
                    break
        except Exception:
            pass
    # Resolution order: explicit image > video (auto-storing the clean
    # zero-detection frame into the reference traf) > frame already stored
    # inside any of the trafs.
    from app.core.background_frame import (resolve_background,
                                           load_stored_background)
    canvas_master = None
    if cfg.get('background_image') and os.path.exists(cfg['background_image']):
        canvas_master = cv2.imread(cfg['background_image'])
        log(f"Trajectory background: image {cfg['background_image']}")
    elif bg_video and os.path.exists(bg_video):
        try:
            canvas_master, desc = resolve_background(
                os.path.abspath(cfg['gates_from']), bg_video, frame='auto')
            log(f"Trajectory background: {desc}")
        except Exception:
            canvas_master = grab_background_frame(bg_video, frame_idx=0)
            log("Trajectory background: video frame 0")
    else:
        for tp in traf_files:
            canvas_master = load_stored_background(tp)
            if canvas_master is not None:
                log(f"Trajectory background: clean frame stored in "
                    f"{os.path.basename(tp)}")
                break
    if canvas_master is None:
        log("Trajectory plots SKIPPED: no background available. Provide "
            "'background_video' or 'background_image' once — the clean "
            "frame is then stored inside the .traf for all future runs.")
        return []
    ch, cw = canvas_master.shape[:2]

    allowed = cfg.get('trajectory_classes') or None
    per_class_cap = cfg.get('trajectory_per_class')
    min_points = int(cfg.get('trajectory_min_points', 10))

    outputs = []
    for p_start, p_end in periods:
        merged = []
        counts = {}
        id_offset = 0
        for traf_path in traf_files:
            conn = sqlite3.connect(traf_path)
            try:
                meta = read_scene_meta(conn)
                try:
                    vs = datetime.fromisoformat(meta['video_start_time'])
                    fps = float(meta.get('fps', 30.0))
                    total = int(meta.get('total_frames', 0))
                except (KeyError, ValueError):
                    log(f"  trajectory: {os.path.basename(traf_path)} has no "
                        f"video_start_time — included without time filter")
                    vs = fps = total = None

                lo = hi = None
                if p_start and vs is not None:
                    def _to_frame(hhmm):
                        h, m = (int(x) for x in hhmm.split(':'))
                        tgt = vs.replace(hour=h, minute=m, second=0)
                        return (tgt - vs).total_seconds() * fps
                    lo, hi = _to_frame(p_start), _to_frame(p_end)
                    if hi <= 0 or (total and lo >= total):
                        conn.close()
                        continue  # file entirely outside this period

                trajs = read_trajectories(
                    conn, min_points=min_points, allowed_classes=allowed,
                    skip_stationary=True, per_class_cap=None)

                if lo is not None:
                    spans = dict(conn.execute(
                        "SELECT track_id, first_frame || ',' || last_frame "
                        "FROM tracks"))
                    def _ok(tid):
                        s = spans.get(tid) or spans.get(int(tid))
                        if not s:
                            return True
                        f0, f1 = (int(x) for x in s.split(','))
                        return f1 >= lo and f0 <= hi
                    trajs = [t for t in trajs if _ok(t[0])]

                # Rescale if this file's frame size differs from the canvas
                fw = int(meta.get('frame_width', cw) or cw)
                fh = int(meta.get('frame_height', ch) or ch)
                sx, sy = cw / fw, ch / fh
                for tid, cls, pts in trajs:
                    if sx != 1.0 or sy != 1.0:
                        pts = pts * np.array([sx, sy], dtype=np.float32)
                    merged.append((id_offset + int(tid), cls, pts))
                    counts[cls] = counts.get(cls, 0) + 1
                id_offset += 1000000
            finally:
                conn.close()

        label = f"{p_start}-{p_end}" if p_start else "full"
        if not merged:
            log(f"Trajectory plot [{label}]: no tracks in window — skipped")
            continue

        # Site-level per-class cap (after merging, so it's fair across files)
        if per_class_cap:
            rng = random.Random(42)
            by_cls = {}
            for t in merged:
                by_cls.setdefault(t[1], []).append(t)
            merged = []
            for cls, items in by_cls.items():
                if len(items) > per_class_cap:
                    items = rng.sample(items, per_class_cap)
                merged.extend(items)

        canvas = canvas_master.copy()
        trajs = despike_trajectories(merged)
        trajs = smooth_trajectories(trajs, sigma=2.0)
        trajs = trim_departing_starts(trajs, frame_h=ch, n_trim=6)
        trajs = filter_noisy_in_crowded_approaches(trajs)
        canvas, drawn = draw_trajectories(canvas, trajs, 1, gradient=True)
        canvas = draw_legend(canvas, counts)
        canvas = draw_job_info(canvas, job_number=cfg.get('job_number'),
                               site_name=site_label)

        suffix = label.replace(':', '')
        out_path = os.path.join(
            out_dir, f"{site_label.replace(' ', '_')}_trajectories_{suffix}.png")
        cv2.imwrite(out_path, canvas)
        outputs.append(out_path)
        total_tracks = sum(counts.values())
        log(f"Trajectory plot [{label}]: {total_tracks} tracks from "
            f"{len(traf_files)} file(s) → {out_path}")
    return outputs


# ══════════════════════════════════════════════════════════════════════
#  Engine (shared by CLI and viewer UI)
# ══════════════════════════════════════════════════════════════════════

def _clean_path(v):
    """Strip stray quotes that ride along with Windows 'Copy as path'."""
    return str(v).strip().strip('"').strip("'").strip() if v else v


def _parse_cfg_date(v):
    if not v:
        return None
    from datetime import date
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise RuntimeError(f"template_date '{v}' not understood — use YYYY-MM-DD")


def run_from_config(cfg, log=print):
    """Run the whole batch extraction from a config dict (see SITE_CONFIG_KEYS).
    Returns {'outputs': [paths], 'warnings': [...], 'qa': [...]}.
    Raises RuntimeError on fatal problems (never calls sys.exit)."""
    for req in ('gates_from', 'traf_dir'):
        if not cfg.get(req):
            raise RuntimeError(f"Config missing required key: {req}")

    for k in ('gates_from', 'traf_dir', 'out', 'schedule',
              'client_template', 'mapping'):
        if cfg.get(k):
            cfg[k] = _clean_path(cfg[k])

    ref_path = os.path.abspath(cfg['gates_from'])
    traf_dir = os.path.abspath(cfg['traf_dir'])
    out_dir = cfg.get('out') or '.'
    site_label = cfg.get('site') or os.path.basename(
        os.path.dirname(traf_dir + os.sep))
    bin_minutes = int(cfg.get('bin', 15))
    os.makedirs(out_dir, exist_ok=True)

    periods = [parse_period(p) for p in (cfg.get('periods') or [])] or [(None, None)]

    traf_files = sorted(glob.glob(os.path.join(traf_dir, '*.traf')))
    if not traf_files:
        raise RuntimeError(f"no .traf files found in {traf_dir}")

    gates, ref_w, ref_h = read_reference_gates(ref_path)
    log(f"Reference gates ({len(gates)}) from {os.path.basename(ref_path)}: "
        f"{', '.join(g[0] for g in gates)}")
    log(f"Files to process: {len(traf_files)}")

    warnings = []
    qa_files = []
    outputs = []

    per_period_results = [[] for _ in periods]
    per_period_file_stats = [[] for _ in periods]

    # ── Class / movement mapping ──
    client_template = cfg.get('client_template')
    client_fill = None
    if client_template:
        from client_template_filler import (load_mapping, accumulate_client_counts,
                                            fill_template, _norm, _norm_mov)
        if cfg.get('mapping'):
            movement_map, class_map = load_mapping(cfg['mapping'])
        else:
            movement_map = {_norm_mov(k): str(v).strip()
                            for k, v in (cfg.get('movements') or {}).items()}
            class_map = {str(k).strip(): _norm(v)
                         for k, v in (cfg.get('classes') or {}).items()}
        if not class_map:
            log("No explicit class map — relying on auto-matching "
                "(exact names + built-in aliases). Check the resolved "
                "mapping in the fill report below.")

        if cfg.get('site_number'):
            arm_modes = {g[0]: g[7] for g in gates if g[7] and g[7] != 'two'}
            arm_modes.update(cfg.get('arm_modes') or {})
            if arm_modes:
                log("One-way arms: " + ", ".join(
                    f"{a}={m}" for a, m in arm_modes.items()))
            auto_map, rows = generate_client_movements(
                [g[0] for g in gates], cfg['site_number'], arm_modes=arm_modes)
            movement_map = {k.replace('->', '→'): v for k, v in auto_map.items()}
            log("Auto-generated movement numbers (verify against the "
                "client's movement sheet):")
            log(f"  {'No.':<8}{'Letters':<10}{'Gates':<22}Turn")
            for num, letters, gates_lbl, turn in rows:
                log(f"  {num:<8}{letters:<10}{gates_lbl:<22}{turn}")

        client_fill = {'movement_map': movement_map, 'class_map': class_map,
                       'counts': {}}
        log(f"Client template: {os.path.basename(client_template)} "
            f"({len(movement_map)} movement(s), {len(class_map)} class(es) mapped)")

    # ── Per-file processing ──
    for fi, traf_path in enumerate(traf_files, 1):
        name = os.path.basename(traf_path)
        is_ref = os.path.samefile(traf_path, ref_path) if os.path.exists(ref_path) else False
        conn = sqlite3.connect(traf_path)

        meta = {r[0]: r[1] for r in conn.execute("SELECT key, value FROM scene")}
        w, h = int(meta.get('frame_width', 0)), int(meta.get('frame_height', 0))
        if ref_w and w and (w != ref_w or h != ref_h):
            log(f"[{fi}/{len(traf_files)}] SKIP  {name} — frame size {w}x{h} "
                f"differs from reference {ref_w}x{ref_h}; gates would not align.")
            qa_files.append({'name': name, 'dur_min': None, 'tracks': 0,
                             'crossings': 0, 'cpm': None,
                             'skipped': f'frame size {w}x{h} vs {ref_w}x{ref_h}',
                             'video_start': None, 'video_end': None})
            conn.close()
            continue

        if not is_ref:
            apply_gates(conn, gates)
        n_cross = recompute_all_gates(conn)
        movements = auto_generate_movements(conn, include_uturns=True)
        n_tracks = conn.execute(
            "SELECT COUNT(*) FROM tracks WHERE is_stationary=0").fetchone()[0]

        tag = " (reference)" if is_ref else ""
        log(f"[{fi}/{len(traf_files)}] {name}{tag}: "
            f"{n_tracks} moving tracks, {n_cross} crossings")

        # QA per-file stats
        dur_min = None
        video_end = None
        try:
            _fps = float(meta.get('fps', 30.0))
            _tf = int(meta.get('total_frames', 0))
            dur_min = _tf / _fps / 60 if _fps else None
            _vs = datetime.fromisoformat(meta['video_start_time'])
            video_end = (_vs + timedelta(seconds=_tf / _fps)).isoformat()
        except (KeyError, ValueError, ZeroDivisionError):
            pass
        qa_files.append({'name': name, 'dur_min': dur_min, 'tracks': n_tracks,
                         'crossings': n_cross,
                         'cpm': (n_cross / dur_min) if dur_min else None,
                         'video_start': meta.get('video_start_time'),
                         'video_end': video_end})
        _qa_gate_direction_check(conn, warnings, name)

        for pi, (p_start, p_end) in enumerate(periods):
            res = compute_time_binned_movements(
                conn, movements, bin_minutes=bin_minutes,
                range_start=p_start, range_end=p_end)
            if not res['has_real_time']:
                conn.close()
                raise RuntimeError(
                    f"{name} has no video_start_time in scene metadata — "
                    f"cannot align clock-time bins across files.")
            per_period_results[pi].append(res)

        if client_fill is not None:
            from client_template_filler import accumulate_client_counts
            res15 = compute_time_binned_movements(
                conn, movements, bin_minutes=15,
                range_start=None, range_end=None)
            try:
                file_start_dt = datetime.fromisoformat(meta['video_start_time'])
            except (KeyError, ValueError):
                conn.close()
                raise RuntimeError(f"{name} has no valid video_start_time — "
                                   f"required for client template date matching.")
            accumulate_client_counts(client_fill['counts'],
                                     client_fill['movement_map'],
                                     res15, file_start_dt)

        for pi, (p_start, p_end) in enumerate(periods):
            res = per_period_results[pi][-1]
            direct = sum(m['direct_total'] for m in res['movements'])
            inferred = sum(m['inferred_total'] for m in res['movements'])
            per_period_file_stats[pi].append({
                'name': name,
                'video_start': meta.get('video_start_time', '—'),
                'video_end': video_end or '—',
                'tracks': n_tracks, 'crossings': n_cross,
                'direct': direct, 'inferred': inferred,
            })
        conn.close()

    # ── Site-level QA ──
    if cfg.get('schedule'):
        _qa_schedule_reconciliation(cfg['schedule'], traf_files, warnings)
    _qa_period_coverage(periods, qa_files, warnings)
    median_cpm = run_qa(qa_files, warnings)

    # ── Merge + write one workbook per period ──
    for pi, (p_start, p_end) in enumerate(periods):
        merged = merge_results(per_period_results[pi])
        if merged is None or not merged['movements']:
            warnings.append(f"No data for period {p_start or 'full'}"
                            f"{('-' + p_end) if p_end else ''} — workbook skipped.")
            log(f"WARNING: no data for period {p_start or 'full'} — skipped.")
            continue

        if p_start:
            period_label = f"{p_start}-{p_end}"
            suffix = f"_{p_start.replace(':', '')}-{p_end.replace(':', '')}"
        else:
            period_label = f"{merged['period_start']} to {merged['period_end']}"
            suffix = "_full"

        wb = build_merged_workbook(merged, site_label, period_label,
                                   per_period_file_stats[pi])
        add_qa_sheet(wb, qa_files, warnings, median_cpm)
        out_name = (f"{site_label.replace(' ', '_')}_TMC_{bin_minutes}min"
                    f"{suffix}.xlsx")
        out_path = os.path.join(out_dir, out_name)
        wb.save(out_path)
        outputs.append(out_path)

        grand = sum(m['grand_total'] for m in merged['movements'])
        log(f"Wrote {out_path}")
        log(f"  Period: {period_label}  |  Bins: {len(merged['time_bins'])}  |  "
            f"Movements: {len(merged['movements'])}  |  Total counted: {grand}")

    # ── Site-level trajectory plots ──
    if cfg.get('trajectory_plots'):
        try:
            t_outs = generate_site_trajectory_plots(
                cfg, traf_files, periods, out_dir, site_label, log)
            outputs.extend(t_outs)
        except Exception as e:
            warnings.append(f"Trajectory plots failed: {e}")
            log(f"WARNING: trajectory plots failed: {e}")

    # ── Fill the client template ──
    if client_fill is not None:
        from client_template_filler import fill_template, _norm
        # Dynamic class aliases from the reference traf's own full names
        extra_aliases = {}
        try:
            _c = sqlite3.connect(ref_path)
            for nm, full in _c.execute(
                    "SELECT DISTINCT class_name, class_full_name FROM tracks "
                    "WHERE class_full_name IS NOT NULL"):
                if nm and full and _norm(nm) != _norm(full):
                    extra_aliases.setdefault(_norm(nm), []).append(_norm(full))
            _c.close()
        except Exception:
            pass
        stem, ext = os.path.splitext(os.path.basename(client_template))
        out_path = os.path.join(out_dir, f"{stem}_filled{ext}")
        report = fill_template(client_template, out_path,
                               client_fill['movement_map'],
                               client_fill['class_map'],
                               client_fill['counts'],
                               extra_aliases=extra_aliases,
                               ignore_dates=bool(cfg.get('ignore_template_dates')),
                               set_template_date=_parse_cfg_date(cfg.get('template_date')))
        outputs.append(out_path)
        log(f"Client template filled → {out_path}")
        if report.get('dates_rewritten'):
            log(f"Template dates rewritten to "
                f"{_parse_cfg_date(cfg.get('template_date')).strftime('%A %d %B %Y')} "
                f"({report['dates_rewritten']} cells)")
        if report.get('class_mapping'):
            log("Resolved class mapping: " + ", ".join(
                f"{k} → {v}" for k, v in report['class_mapping'].items()))
        for s in report['sheets']:
            log(f"  {s['sheet']} ({s['movement']}): {s['vehicles_placed']} vehicles "
                f"into {s['bins_with_data']} bins "
                f"({s['bins_zeroed']} bins zero-filled)")
        for w in report['warnings']:
            warnings.append(w)
            log(f"  WARNING: {w}")

    if warnings:
        log(f"\nQA: {len(warnings)} warning(s) — see the QA sheet in the workbook:")
        for w in warnings:
            log(f"  ⚠ {w}")
    else:
        log("\nQA: PASS — no issues found.")
    log("Done. The .traf files keep their gates + crossings.")
    return {'outputs': outputs, 'warnings': warnings, 'qa': qa_files}


def main():
    parser = argparse.ArgumentParser(
        description='Batch TMC extraction: copy gates from a reference .traf '
                    'to all files of a location, recompute crossings, and '
                    'produce merged time-binned Excel reports with QA.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Recommended: put all settings in a site YAML and run\n"
               "  python batch_report.py --config site42.yaml\n" + SITE_CONFIG_KEYS)
    parser.add_argument('--config', default=None,
                        help='Site config YAML — any CLI flag below overrides it')
    parser.add_argument('--gates-from', default=None)
    parser.add_argument('--traf-dir', default=None)
    parser.add_argument('--bin', type=int, default=None, choices=[1, 5, 15, 30, 60])
    parser.add_argument('--period', action='append', default=[],
                        metavar='HH:MM-HH:MM')
    parser.add_argument('--out', default=None)
    parser.add_argument('--site', default=None)
    parser.add_argument('--client-template', default=None)
    parser.add_argument('--mapping', default=None)
    parser.add_argument('--auto-movements', action='store_true')
    parser.add_argument('--site-number', default=None)
    parser.add_argument('--arm-mode', action='append', default=[], metavar='ARM=MODE')
    parser.add_argument('--schedule', default=None,
                        help='Video_schedule.xlsx for QA reconciliation')
    parser.add_argument('--trajectory-plots', action='store_true',
                        help='Generate merged site trajectory PNGs per period')
    parser.add_argument('--background-video', default=None,
                        help='Video for the plot background frame '
                             '(default: first reachable video in the schedule)')
    args = parser.parse_args()

    cfg = load_site_config(args.config) if args.config else {}
    # CLI overrides YAML
    if args.gates_from: cfg['gates_from'] = args.gates_from
    if args.traf_dir: cfg['traf_dir'] = args.traf_dir
    if args.bin: cfg['bin'] = args.bin
    if args.period: cfg['periods'] = args.period
    if args.out: cfg['out'] = args.out
    if args.site: cfg['site'] = args.site
    if args.client_template: cfg['client_template'] = args.client_template
    if args.mapping: cfg['mapping'] = args.mapping
    if args.site_number: cfg['site_number'] = args.site_number
    if args.schedule: cfg['schedule'] = args.schedule
    if args.trajectory_plots: cfg['trajectory_plots'] = True
    if args.background_video: cfg['background_video'] = args.background_video
    if args.arm_mode:
        modes = dict(cfg.get('arm_modes') or {})
        for spec in args.arm_mode:
            if '=' not in spec:
                sys.exit(f"ERROR: --arm-mode expects ARM=MODE, got '{spec}'")
            a, m = spec.split('=', 1)
            modes[a] = m
        cfg['arm_modes'] = modes
    if args.auto_movements and not cfg.get('site_number'):
        parser.error('--auto-movements requires --site-number (or site_number in YAML)')

    try:
        run_from_config(cfg)
    except RuntimeError as e:
        sys.exit(f"ERROR: {e}")


if __name__ == '__main__':
    main()
