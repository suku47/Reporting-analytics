"""
client_template_filler.py — Fill a client's per-movement Excel template
(.xlsm/.xlsx) directly from .traf-derived counts.

Template assumptions (verified against 'Basic Data Format - All Sites'):
  - One sheet per movement, named 'Movement <number>' (e.g. 'Movement 2.1')
  - Each sheet contains one or more session blocks; each block has:
      'Date' label row  → date value row ('Monday 06 April 2026')
      ... 'TIME' header row with client class columns ...
      15-minute bin rows labelled 'HHMM - HHMM' (e.g. '0700 - 0715')
      'Hourly Total'/'Hourly Average'/'Session ...' rows (formulas — untouched)
  - Only raw counts are written into bin-row × mapped-class cells;
    every total/average is already a formula and recalculates in Excel.

Mapping JSON:
{
  "movements": { "C→D": "2.1", "C→B": "2.2", "C→C": "2.3" },
  "classes":   { "Car": "Cars (LMV)", "Bus": "BMTC Buses" }
}
Movement keys accept '→' or '->'. Class keys are OUR detector class names;
values are the client's column headers (matched case-insensitively,
whitespace-normalised).

Fill policy (user-confirmed):
  - Bins with no data from our recordings are written as 0 (mapped columns only)
  - Rows are matched by DATE + TIME strictly (template date row must equal
    the survey date derived from the .traf video_start_time)
"""

import json
import re
import sys
from datetime import datetime, timedelta

BIN_ROW_RE = re.compile(r'^(\d{2})(\d{2})\s*-\s*(\d{2})(\d{2})$')
DATE_FORMATS = ('%A %d %B %Y', '%d %B %Y', '%A %d %b %Y', '%d %b %Y',
                '%d/%m/%Y', '%Y-%m-%d')

# Known-equivalent names between our detector classes and common client
# column headers (all compared lowercase, whitespace-collapsed). Explicit
# entries in the YAML 'classes:' section always win over these.
DEFAULT_CLASS_ALIASES = {
    'car':        ['car', 'cars'],
    'taxi':       ['taxi', 'taxis'],
    'lgv':        ['lgv', 'lgvs'],
    'ogv1':       ['ogv1', 'ogv 1'],
    'ogv2':       ['ogv2', 'ogv 2'],
    'bus':        ['bus/coach', 'bus / coach', 'bus', 'buses', 'bus & coach',
                   'buses/coaches'],
    'biker':      ['m/cycle', 'm / cycle', 'motorcycle', 'motor cycle',
                   'motorcycles', 'mcycle'],
    'motorcycle': ['m/cycle', 'm / cycle', 'motorcycle', 'motorcycles'],
    'cyclist':    ['p/cycle', 'p / cycle', 'pedal cycle', 'pedal cycles',
                   'cycle', 'bicycle', 'bicycles'],
    'pedal cycle': ['p/cycle', 'p / cycle', 'pedal cycle'],
    'minibus':    ['minibus', 'mini bus'],
    # US (FHWA simplified) profile
    'pv':         ['pv', 'passenger vehicle', 'passenger vehicles', 'passenger cars'],
    'su':         ['su', 'single-unit truck', 'single unit truck',
                   'single-unit trucks', 'single unit trucks', 'su truck'],
    'cu':         ['cu', 'combination-unit truck', 'combination unit truck',
                   'combination-unit trucks', 'combination trucks',
                   'articulated truck', 'articulated trucks'],
}


def auto_match_class(our_class, header_cols, extra_aliases=None):
    """Find the header column for one of our classes: exact name match
    first, then the static alias table, then dynamic aliases (e.g. the
    class_full_name stored in the .traf itself, so ANY class profile —
    UK, US, India, future — matches without code changes).
    header_cols: {normalised header: col}. Returns the header key or None."""
    key = _norm(our_class)
    if key in header_cols:
        return key
    candidates = list(DEFAULT_CLASS_ALIASES.get(key, []))
    for c in (extra_aliases or {}).get(key, []):
        candidates.append(c)
        candidates.append(c + 's')          # simple plural
        if c.endswith('s'):
            candidates.append(c[:-1])       # simple singular
    for candidate in candidates:
        if candidate in header_cols:
            return candidate
    return None


def _norm(s):
    """Normalise header text for matching: lowercase, collapse whitespace."""
    return re.sub(r'\s+', ' ', str(s)).strip().lower()


def _norm_mov(key):
    """Normalise movement key: 'C -> D' / 'C→D' → 'C→D'."""
    key = key.replace('->', '→')
    parts = [p.strip() for p in key.split('→')]
    return '→'.join(parts)


def load_mapping(path):
    with open(path, encoding='utf-8') as f:
        raw = json.load(f)
    if 'classes' not in raw:
        sys.exit(f"ERROR: mapping file {path} must contain a 'classes' key")
    movements = {_norm_mov(k): str(v).strip()
                 for k, v in raw.get('movements', {}).items()}
    classes = {str(k).strip(): _norm(v) for k, v in raw['classes'].items()}
    return movements, classes


def _parse_template_date(value):
    if isinstance(value, datetime):
        return value.date()
    if value is None:
        return None
    s = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def accumulate_client_counts(store, movements, res, file_start_dt):
    """Accumulate one file's full-recording 15-min result into `store`:
    store[mov_label][(date, 'HH:MM')][our_class] += n

    A bin's date is the file's start date; if the bin's start time-of-day is
    earlier than the file's start time (midnight wrap), it belongs to the
    next day.
    """
    file_date = file_start_dt.date()
    file_tod = file_start_dt.time()

    for mov in res['movements']:
        label = _norm_mov(f"{mov['from_name']}→{mov['to_name']}")
        mov_store = store.setdefault(label, {})
        for bi, bin_label in enumerate(res['time_bins']):
            m = re.match(r'^(\d{1,2}):(\d{2})', bin_label)
            if not m:
                continue
            hh, mm = int(m.group(1)), int(m.group(2))
            bin_date = file_date
            if (hh, mm) < (file_tod.hour, file_tod.minute) and file_tod.hour >= 12:
                bin_date = file_date + timedelta(days=1)  # midnight wrap
            key = (bin_date, f"{hh:02d}:{mm:02d}")
            cell = mov_store.setdefault(key, {})
            for cls, n in mov['bins'][bi].items():
                if n:
                    cell[cls] = cell.get(cls, 0) + n


def fill_template(template_path, out_path, movement_map, class_map,
                  client_counts, extra_aliases=None, ignore_dates=False,
                  set_template_date=None):
    """set_template_date: a datetime.date — every date cell in every
    movement sheet is REWRITTEN to this date (correct weekday computed
    automatically) before filling. Kills the edit-18-cells-by-hand chore
    when reusing a template from an older job."""
    """Fill the client workbook. Returns a report dict."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl required")

    keep_vba = template_path.lower().endswith('.xlsm')
    wb = openpyxl.load_workbook(template_path, keep_vba=keep_vba)

    report = {'sheets': [], 'warnings': []}

    # Classes present in the data: explicit map entries win, the rest are
    # auto-matched per sheet against the template headers (alias table).
    seen_classes = set()
    for mov_store in client_counts.values():
        for cell in mov_store.values():
            seen_classes.update(cell.keys())
    ever_matched = set()          # our classes that found a column somewhere
    resolved_pairs = {}           # our class -> header actually used (report)

    # Survey dates present in the data (for the date-mismatch check) and,
    # for ignore_dates mode, per-movement counts re-keyed by time-of-day.
    data_dates = set()
    time_only_counts = {}
    for mov_key, mov_store in client_counts.items():
        t_store = time_only_counts.setdefault(mov_key, {})
        for (d, t), cell in mov_store.items():
            data_dates.add(d)
            tgt = t_store.setdefault(t, {})
            for cls, n in cell.items():
                tgt[cls] = tgt.get(cls, 0) + n
    template_dates = set()
    dates_rewritten = 0

    for mov_key, mov_number in movement_map.items():
        sheet_name = f"Movement {mov_number}"
        if sheet_name not in wb.sheetnames:
            report['warnings'].append(
                f"Sheet '{sheet_name}' (for {mov_key}) not found in template — skipped. "
                f"Available: {', '.join(n for n in wb.sheetnames if n.startswith('Movement'))[:200]}")
            continue

        ws = wb[sheet_name]
        mov_counts = client_counts.get(mov_key, {})

        current_date = None
        header_cols = {}          # normalised header → column index
        cells_written = 0
        bins_with_data = 0
        bins_zeroed = 0
        vehicles_placed = 0
        expect_date_value = False
        sheet_dates = set()        # dates seen on this sheet
        pending_bin_rows = []      # bin row indices since last Hourly Total
        last_hourly_sums = None    # {col: (total, n_bins)} for the Average row
        effective_map = {}         # our class -> header (explicit + auto)

        for row in ws.iter_rows(min_col=1, max_col=ws.max_column):
            a_val = row[0].value
            a_str = str(a_val).strip() if a_val is not None else ''

            if a_str.lower() == 'date':
                expect_date_value = True
                continue
            if expect_date_value:
                d = _parse_template_date(a_val)
                if d or (set_template_date and a_str):
                    if set_template_date:
                        d = set_template_date
                        row[0].value = d.strftime('%A %d %B %Y')
                        dates_rewritten += 1
                    current_date = d
                    sheet_dates.add(d)
                    expect_date_value = False
                    continue
                # not a date value → keep looking on subsequent rows
                if a_str:
                    expect_date_value = False
                continue

            if a_str.upper() == 'TIME':
                header_cols = {}
                for cell in row[1:]:
                    if cell.value:
                        header_cols[_norm(cell.value)] = cell.column
                # Effective map for this block: explicit entries first,
                # auto-matched (exact name or alias) for the rest.
                effective_map = {}
                for our_cls, client_hdr in class_map.items():
                    if client_hdr in header_cols:
                        effective_map[our_cls] = client_hdr
                for our_cls in seen_classes:
                    if our_cls in effective_map:
                        continue
                    hdr = auto_match_class(our_cls, header_cols, extra_aliases)
                    if hdr is not None:
                        effective_map[our_cls] = hdr
                for our_cls, hdr in effective_map.items():
                    ever_matched.add(our_cls)
                    resolved_pairs.setdefault(our_cls, hdr)
                continue

            m = BIN_ROW_RE.match(a_str)

            # Hourly Total / Hourly Average rows hold LITERAL values in the
            # client template (only Session rows are formulas), so we must
            # write computed hourlies for mapped columns or they'd stay stale.
            if a_str.lower() == 'hourly total' and pending_bin_rows and header_cols:
                hourly_sums = {}
                for our_cls, client_hdr in effective_map.items():
                    col = header_cols.get(client_hdr)
                    if col is None:
                        continue
                    total = sum((ws.cell(row=r, column=col).value or 0)
                                for r in pending_bin_rows)
                    hourly_sums[col] = (total, len(pending_bin_rows))
                    ws.cell(row=row[0].row, column=col, value=total)
                    cells_written += 1
                last_hourly_sums = hourly_sums
                pending_bin_rows = []
                continue

            if a_str.lower() == 'hourly average' and last_hourly_sums:
                for col, (total, n_bins) in last_hourly_sums.items():
                    ws.cell(row=row[0].row, column=col,
                            value=round(total / n_bins, 2) if n_bins else 0)
                    cells_written += 1
                last_hourly_sums = None
                continue

            if not (m and header_cols and current_date):
                continue

            if current_date:
                template_dates.add(current_date)
            t_key = f"{m.group(1)}:{m.group(2)}"
            if ignore_dates:
                counts = time_only_counts.get(mov_key, {}).get(t_key, {})
            else:
                counts = mov_counts.get(bin_key := (current_date, t_key), {})
            row_has_data = bool(counts)
            pending_bin_rows.append(row[0].row)

            for our_cls, client_hdr in effective_map.items():
                col = header_cols.get(client_hdr)
                if col is None:
                    continue
                n = counts.get(our_cls, 0)
                ws.cell(row=row[0].row, column=col, value=n)
                cells_written += 1
                vehicles_placed += n

            if row_has_data:
                bins_with_data += 1
            else:
                bins_zeroed += 1

        # Explicitly-mapped headers that never matched a column on this sheet
        missing_hdrs = [h for h in set(class_map.values())
                        if header_cols and h not in header_cols]
        if missing_hdrs:
            report['warnings'].append(
                f"{sheet_name}: mapped column header(s) not found: "
                f"{', '.join(missing_hdrs)}")

        if (not ignore_dates and not set_template_date and data_dates
                and sheet_dates and not (data_dates & sheet_dates)):
            report['warnings'].append(
                f"{sheet_name}: sheet dates ("
                + ", ".join(sorted(d.isoformat() for d in sheet_dates))
                + ") don't match any survey date — its bins were zero-filled.")

        report['sheets'].append({
            'sheet': sheet_name, 'movement': mov_key,
            'cells_written': cells_written,
            'bins_with_data': bins_with_data, 'bins_zeroed': bins_zeroed,
            'vehicles_placed': vehicles_placed,
        })

    if dates_rewritten:
        report['dates_rewritten'] = dates_rewritten
    if resolved_pairs:
        report['class_mapping'] = dict(sorted(resolved_pairs.items()))

    # Loud warning when strict matching can never fill anything
    if (not ignore_dates and data_dates and template_dates
            and not (data_dates & template_dates)):
        report['warnings'].append(
            "TEMPLATE DATE MISMATCH: template rows are dated "
            + ", ".join(sorted(d.isoformat() for d in template_dates))
            + " but the survey data is "
            + ", ".join(sorted(d.isoformat() for d in data_dates))
            + " — every bin was zero-filled. Fix the dates in the client "
            "template, or set ignore_template_dates: true to match by "
            "time-of-day only.")
    never = sorted(c for c in seen_classes if c not in ever_matched)
    if never:
        report['warnings'].append(
            f"Classes present in data but matched to NO template column "
            f"(counts NOT written): {', '.join(never)}. Add them to "
            f"'classes:' in the site YAML if the client needs them.")

    wb.save(out_path)
    report['out_path'] = out_path
    return report


def print_report(report):
    print(f"\nClient template filled → {report['out_path']}")
    for s in report['sheets']:
        print(f"  {s['sheet']}  ({s['movement']}): "
              f"{s['vehicles_placed']} vehicles into {s['bins_with_data']} bins "
              f"({s['bins_zeroed']} bins zero-filled, {s['cells_written']} cells written)")
    for w in report['warnings']:
        print(f"  WARNING: {w}")
