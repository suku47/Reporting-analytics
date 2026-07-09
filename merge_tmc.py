"""
merge_tmc.py  — Merge multiple TMC xlsx files into one consolidated workbook.
Handles: Time-Binned Counts, all 6 direction sheets (A→B etc.), Classification Summary.
Overlapping time bins are SUMMED across files.

Usage:
    python merge_tmc.py                          # default: /mnt/user-data/uploads/VID_*.xlsx
    python merge_tmc.py -i /path/to/folder -o merged.xlsx
    python merge_tmc.py -i file1.xlsx file2.xlsx -o merged.xlsx
"""

import argparse, glob, os, sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── constants ─────────────────────────────────────────────────────────────────
MOVEMENTS   = ["A → B", "A → C", "B → A", "B → C", "C → A", "C → B"]
SHEET_DIRS  = ["A→B", "A→C", "B→A", "B→C", "C→A", "C→B"]   # as they appear in xlsx
DIR_CLASSES = ["Bus", "Car", "LGV", "OGV1"]
CLASS_COLS  = ["Bus", "Car", "LGV", "OGV1"]        # classification summary
EXTRA_COLS  = ["Direct", "Inferred"]

SURVEY_PERIODS = [("08:00", "10:00"), ("15:00", "17:00")]

# ── style helpers ──────────────────────────────────────────────────────────────
def _fill(hex_): return PatternFill("solid", start_color=hex_)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, name="Arial", italic=italic)
thin  = Side(border_style="thin", color="BBBBBB")
BORD  = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR   = Alignment(horizontal="center", vertical="center")
LEFT  = Alignment(horizontal="left",   vertical="center")

DARK_BLUE  = "1F4E79"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
ALT_BLUE   = "EBF3FB"
WHITE      = "FFFFFF"

def _hdr(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(bold=bold, color=fg); c.fill = _fill(bg)
    c.alignment = CTR; c.border = BORD; return c

def _data(ws, row, col, val, bg=None, align=CTR):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _font(); c.fill = _fill(bg) if bg else PatternFill("none")
    c.alignment = align; c.border = BORD; return c

def _total_row(ws, row, ncols, label, data_start, data_end, bg=LIGHT_BLUE):
    c = ws.cell(row=row, column=1, value=label)
    c.font = _font(bold=True); c.fill = _fill(bg); c.alignment = LEFT; c.border = BORD
    for col in range(2, ncols + 1):
        cl = get_column_letter(col)
        cell = ws.cell(row=row, column=col,
                       value=f"=SUM({cl}{data_start}:{cl}{data_end})")
        cell.font = _font(bold=True); cell.fill = _fill(bg)
        cell.alignment = CTR; cell.border = BORD

def _section_hdr(ws, row, ncols, label):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=label)
    c.font = _font(bold=True, color=WHITE); c.fill = _fill(MED_BLUE)
    c.alignment = LEFT; return c

def _title_block(ws, ncols, title, subtitle, source_line):
    last = get_column_letter(ncols)
    ws.merge_cells(f"A1:{last}1")
    c = ws["A1"]; c.value = title
    c.font = Font(bold=True, name="Arial", size=12, color=DARK_BLUE); c.alignment = CTR
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{last}2")
    c = ws["A2"]; c.value = subtitle
    c.font = _font(italic=True, color="555555", size=9); c.alignment = CTR

    ws.merge_cells(f"A3:{last}3")
    c = ws["A3"]; c.value = source_line
    c.font = _font(italic=True, color="777777", size=8); c.alignment = CTR

# ── expected time bins ─────────────────────────────────────────────────────────
def build_expected_bins(periods=SURVEY_PERIODS):
    bins = []
    for sh, eh in periods:
        t = int(sh[:2])*60 + int(sh[3:])
        e = int(eh[:2])*60 + int(eh[3:])
        while t < e:
            s = f"{t//60:02d}:{t%60:02d}"; t += 15
            bins.append(f"{s} - {t//60:02d}:{t%60:02d}")
    return bins

EXPECTED_BINS = build_expected_bins()

# ── sheet parsers ──────────────────────────────────────────────────────────────
def _find_header_row(df, marker):
    for i, row in df.iterrows():
        if any(str(v).strip() == marker for v in row):
            return i
    return None

def parse_time_binned(filepath):
    """Returns {time_bin: {movement: count}}"""
    df = pd.read_excel(filepath, sheet_name="Time-Binned Counts", header=None)
    hr = _find_header_row(df, "Time Bin")
    if hr is None: return {}
    data = pd.read_excel(filepath, sheet_name="Time-Binned Counts", header=hr)
    data.columns = [str(c).strip() for c in data.columns]
    # normalise arrow spacing
    data.rename(columns=lambda x: x.replace("→"," → ").replace("  "," ").strip(), inplace=True)
    data = data[data["Time Bin"].notna()]
    data = data[~data["Time Bin"].astype(str).str.upper().str.contains("TOTAL")]
    out = {}
    for _, row in data.iterrows():
        tb = str(row["Time Bin"]).strip()
        out[tb] = {}
        for m in MOVEMENTS:
            if m in row and pd.notna(row[m]):
                out[tb][m] = int(row[m])
    return out

def parse_direction_sheet(filepath, sheet_name):
    """Returns {time_bin: {class: count}} for one direction sheet."""
    xl = pd.ExcelFile(filepath)
    if sheet_name not in xl.sheet_names:
        return {}
    df = pd.read_excel(filepath, sheet_name=sheet_name, header=None)
    hr = _find_header_row(df, "Time Bin")
    if hr is None: return {}
    data = pd.read_excel(filepath, sheet_name=sheet_name, header=hr)
    data.columns = [str(c).strip() for c in data.columns]
    data = data[data["Time Bin"].notna()]
    data = data[~data["Time Bin"].astype(str).str.upper().str.contains("TOTAL")]
    out = {}
    for _, row in data.iterrows():
        tb = str(row["Time Bin"]).strip()
        out[tb] = {}
        for cls in DIR_CLASSES:
            if cls in row and pd.notna(row[cls]):
                out[tb][cls] = int(row[cls])
    return out

def parse_classification_summary(filepath):
    """Returns {movement: {Bus, Car, LGV, OGV1, Direct, Inferred}}"""
    xl = pd.ExcelFile(filepath)
    if "Classification Summary" not in xl.sheet_names:
        return {}
    df = pd.read_excel(filepath, sheet_name="Classification Summary", header=None)
    hr = _find_header_row(df, "Movement")
    if hr is None: return {}
    data = pd.read_excel(filepath, sheet_name="Classification Summary", header=hr)
    data.columns = [str(c).strip() for c in data.columns]
    data = data[data["Movement"].notna()]
    data = data[~data["Movement"].astype(str).str.upper().str.contains("GRAND TOTAL")]
    out = {}
    for _, row in data.iterrows():
        mv = str(row["Movement"]).strip()
        # normalise arrow
        mv = mv.replace("→", " → ").replace("  ", " ")
        out[mv] = {}
        for col in CLASS_COLS + EXTRA_COLS:
            if col in row and pd.notna(row[col]):
                out[mv][col] = int(row[col])
    return out

# ── merge logic ────────────────────────────────────────────────────────────────
def merge_all(files):
    # Time-binned totals: {time_bin: {movement: count}}
    tb_totals = defaultdict(lambda: defaultdict(int))
    # Direction totals: {sheet_name: {time_bin: {class: count}}}
    dir_totals = {s: defaultdict(lambda: defaultdict(int)) for s in SHEET_DIRS}
    # Classification totals: {movement: {col: count}}
    cls_totals = defaultdict(lambda: defaultdict(int))

    for f in sorted(files):
        # Time-Binned Counts
        for tb, mvs in parse_time_binned(f).items():
            for mv, cnt in mvs.items():
                tb_totals[tb][mv] += cnt

        # Direction sheets
        for sname in SHEET_DIRS:
            for tb, classes in parse_direction_sheet(f, sname).items():
                for cls, cnt in classes.items():
                    dir_totals[sname][tb][cls] += cnt

        # Classification Summary
        for mv, cols in parse_classification_summary(f).items():
            for col, cnt in cols.items():
                cls_totals[mv][col] += cnt

    return tb_totals, dir_totals, cls_totals

# ── sheet writers ──────────────────────────────────────────────────────────────
def write_time_binned(wb, tb_totals, source_files):
    ws = wb.active
    ws.title = "Consolidated TMC"
    ncols = len(MOVEMENTS) + 2  # Time Bin + 6 movements + Total

    _title_block(ws, ncols,
        "Traffic Movement Counts — Consolidated",
        "Survey Date: 15 May 2026  |  Periods: 08:00–10:00 & 15:00–17:00  |  Bin: 15 min",
        f"Merged from {len(source_files)} file(s): {', '.join(Path(f).name for f in sorted(source_files))}")

    # header row
    headers = ["Time Bin"] + MOVEMENTS + ["Total"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 5, ci, h)
    ws.row_dimensions[5].height = 18

    sections = [
        ("08:00 – 10:00 (Morning Peak)",    [b for b in EXPECTED_BINS if b[:2] in ("08","09")]),
        ("15:00 – 17:00 (Afternoon Peak)",  [b for b in EXPECTED_BINS if b[:2] in ("15","16")]),
    ]

    cur = 6
    for sec_label, bins in sections:
        if not bins: continue
        _section_hdr(ws, cur, ncols, sec_label); cur += 1
        ds = cur
        for i, tb in enumerate(bins):
            bg = ALT_BLUE if i % 2 else None
            counts = tb_totals.get(tb, {})
            vals = [counts.get(m, 0) for m in MOVEMENTS]
            total = sum(vals)
            _data(ws, cur, 1, tb, bg, LEFT)
            for ci, v in enumerate(vals, 2):
                _data(ws, cur, ci, v, bg)
            _data(ws, cur, len(MOVEMENTS)+2, total, bg)
            cur += 1
        de = cur - 1
        _total_row(ws, cur, ncols, "Sub-Total", ds, de); cur += 2

    # Grand total
    _hdr(ws, cur, 1, "GRAND TOTAL", DARK_BLUE, WHITE)
    for ci in range(2, ncols + 1):
        cl = get_column_letter(ci)
        cell = ws.cell(row=cur, column=ci,
                       value=f"=SUM({cl}6:{cl}{cur-1})")
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _fill(DARK_BLUE); cell.alignment = CTR; cell.border = BORD

    ws.column_dimensions["A"].width = 18
    for ci in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "B6"


def write_direction_sheet(wb, sheet_name, dir_data, source_files):
    """One sheet per direction (e.g. A→B) with time bins × vehicle classes."""
    ws = wb.create_sheet(sheet_name)
    # display name like "A → B"
    display = sheet_name.replace("→", " → ")
    ncols = len(DIR_CLASSES) + 2  # Time Bin + classes + Total

    _title_block(ws, ncols,
        f"Movement: {display} — Consolidated Classification",
        "Survey Date: 15 May 2026  |  Periods: 08:00–10:00 & 15:00–17:00  |  Bin: 15 min",
        f"Merged from {len(source_files)} file(s): {', '.join(Path(f).name for f in sorted(source_files))}")

    headers = ["Time Bin"] + DIR_CLASSES + ["Total"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 5, ci, h)
    ws.row_dimensions[5].height = 18

    sections = [
        ("08:00 – 10:00 (Morning Peak)",   [b for b in EXPECTED_BINS if b[:2] in ("08","09")]),
        ("15:00 – 17:00 (Afternoon Peak)", [b for b in EXPECTED_BINS if b[:2] in ("15","16")]),
    ]

    cur = 6
    for sec_label, bins in sections:
        if not bins: continue
        _section_hdr(ws, cur, ncols, sec_label); cur += 1
        ds = cur
        for i, tb in enumerate(bins):
            bg = ALT_BLUE if i % 2 else None
            counts = dir_data.get(tb, {})
            vals = [counts.get(cls, 0) for cls in DIR_CLASSES]
            total = sum(vals)
            _data(ws, cur, 1, tb, bg, LEFT)
            for ci2, v in enumerate(vals, 2):
                _data(ws, cur, ci2, v, bg)
            _data(ws, cur, len(DIR_CLASSES)+2, total, bg)
            cur += 1
        de = cur - 1
        _total_row(ws, cur, ncols, "Sub-Total", ds, de); cur += 2

    _hdr(ws, cur, 1, "GRAND TOTAL", DARK_BLUE, WHITE)
    for ci in range(2, ncols + 1):
        cl = get_column_letter(ci)
        cell = ws.cell(row=cur, column=ci,
                       value=f"=SUM({cl}6:{cl}{cur-1})")
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _fill(DARK_BLUE); cell.alignment = CTR; cell.border = BORD

    ws.column_dimensions["A"].width = 18
    for ci in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.freeze_panes = "B6"


def write_classification_summary(wb, cls_totals, source_files):
    ws = wb.create_sheet("Classification Summary")
    all_cols = CLASS_COLS + EXTRA_COLS
    ncols = 1 + len(all_cols) + 1  # Movement + classes + Direct + Inferred + Total

    _title_block(ws, ncols,
        "Vehicle Classification Summary — Consolidated (All Bins)",
        "Survey Date: 15 May 2026  |  Periods: 08:00–10:00 & 15:00–17:00",
        f"Merged from {len(source_files)} file(s): {', '.join(Path(f).name for f in sorted(source_files))}")

    headers = ["Movement"] + all_cols + ["Total"]
    for ci, h in enumerate(headers, 1):
        _hdr(ws, 5, ci, h)
    ws.row_dimensions[5].height = 18

    # movements in a consistent order: use MOVEMENTS list normalised
    ordered = [m.replace(" ", "") for m in MOVEMENTS]  # "A→B" etc
    cur = 6
    ds = cur
    for i, mv_norm in enumerate(ordered):
        # find matching key in cls_totals (may have spaces or not)
        matched = None
        for k in cls_totals:
            if k.replace(" ", "") == mv_norm:
                matched = k; break
        counts = cls_totals.get(matched, {}) if matched else {}
        bg = ALT_BLUE if i % 2 else None
        display = mv_norm.replace("→", " → ")
        _data(ws, cur, 1, display, bg, LEFT)
        row_vals = [counts.get(c, 0) for c in all_cols]
        total = sum(counts.get(c, 0) for c in CLASS_COLS)  # total vehicles = sum of classes
        for ci, v in enumerate(row_vals, 2):
            _data(ws, cur, ci, v, bg)
        _data(ws, cur, ncols, total, bg)
        cur += 1
    de = cur - 1
    _total_row(ws, cur, ncols, "GRAND TOTAL", ds, de, DARK_BLUE)
    # fix font colour on total row
    for ci in range(1, ncols + 1):
        cell = ws.cell(row=cur, column=ci)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _fill(DARK_BLUE)

    ws.column_dimensions["A"].width = 14
    for ci in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.freeze_panes = "B6"


# ── main ───────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Merge TMC xlsx files")
    parser.add_argument("-i", "--input", nargs="+", default=None)
    parser.add_argument("-o", "--output", default="merged_tmc.xlsx")
    args = parser.parse_args()

    if args.input is None:
        print(f"None: {args.input}")
        files = sorted(glob.glob("/mnt/user-data/uploads/VID_*.xlsx"))
    elif len(args.input) == 1 and os.path.isdir(args.input[0]):
        print(f"file: {args.input}")
        files = sorted(glob.glob(os.path.join(args.input[0], "VID_*.xlsx")))
    else:
        print(f"Using input files: {args.input}")
        files = args.input

    if not files:
        print("No input files found.", file=sys.stderr); sys.exit(1)

    print(f"Processing {len(files)} file(s):")
    for f in files: print(f"  {Path(f).name}")

    tb_totals, dir_totals, cls_totals = merge_all(files)

    wb = Workbook()

    print("\nWriting: Consolidated TMC (Time-Binned Counts)...")
    write_time_binned(wb, tb_totals, files)

    for sname in SHEET_DIRS:
        print(f"Writing: {sname}...")
        write_direction_sheet(wb, sname, dir_totals[sname], files)

    print("Writing: Classification Summary...")
    write_classification_summary(wb, cls_totals, files)

    wb.save(args.output)
    print(f"\nSaved → {args.output}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")

if __name__ == "__main__":
    main()
